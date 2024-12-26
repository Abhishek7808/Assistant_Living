import os
import sys

from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from xlsxwriter import Workbook, worksheet
from selenium import webdriver
import pandas as pd
import pyodbc
import glob
import os.path
import stat
from xlsxwriter.worksheet import Worksheet

sys.path[:0] = ['../../../../']
sys.path[:0] = ['../../']
from pythonframework.ProjectTemplate.CommonMethodLibrary.Appium_Android import Appium_Android
from pythonframework.ProjectTemplate.CommonMethodLibrary.SeleniumCommonLibrary import SeleniumCommonLibrary


import pandas as pd
from pythonframework.FrameworkCore.ReadExcelData.ReadExcel import ReadExcel
from pythonframework.FrameworkCore.ObjectRepository.ObjectRepository import TestObject, SubObject
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import ConfigFile
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import Control_TestCase_File
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import AutomationApproach
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import ORFormat
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import ORApplicable
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import ScriptlessOR
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import ScriptlessORKeywords
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import TestCaseResults
from pythonframework.FrameworkCore.FrameworkLibrary.FrameworkCommonLibrary import FrameworkCommonLibrary
from pythonframework.FrameworkCore.CoreLibrary.SeleniumCoreLibrary import SeleniumCoreLibrary
from pythonframework.FrameworkCore.CoreLibrary.AppiumCoreLibrary import AppiumCoreLibrary
from pythonframework.FrameworkCore.TestResultReport.ColorPrint import ColorPrint
from pythonframework.FrameworkCore.SeleniumConfiguration.ConfigurationManager import ConfigurationManager
import time
from datetime import datetime

class test_check:

    def __init__(self, FrameworkInitializer):
        self.frameworkInitializer = FrameworkInitializer
        self.Driver = FrameworkInitializer.Driver
        self.AppiumDriver = FrameworkInitializer.AppiumDriver
        self.AppiumAndroidDriver = FrameworkInitializer.AppiumAndroidDriver
        self.AppiumIosDriver = FrameworkInitializer.AppiumIosDriver
        self.AppiumServer = None
        self.ObjTestResult = FrameworkInitializer.ObjTestResult
        self.FrameworkInitializer = FrameworkInitializer
        self.ObjSeleniumConfiguration = FrameworkInitializer.ObjSeleniumConfiguration
        self.FrameworkCommonLibrary = FrameworkInitializer.ObjFrameworkCommon
        self.ObjConfigurationManager = FrameworkInitializer.ObjConfigurationManager
        self.Appium_Android = Appium_Android(FrameworkInitializer)
        self.ObjSeleniumCommonLib = SeleniumCommonLibrary(FrameworkInitializer)
        self.ObjectRepository = FrameworkInitializer.ObjectRepository




    #for web application
    def webappexample(self):
        print("*****************Vivek*****************")
        self.ObjSeleniumCommonLib.OpenApplication("https://172.16.231.61:447/")
        self.ObjSeleniumCommonLib.ScreenCapture("beforelogin")
        self.ObjSeleniumCommonLib.SetValueInEditBox("karan.agarwal@securemeters.com","BBcoolLogin","Email")
        self.ObjSeleniumCommonLib.SetValueInEditBox("secure@123","BBcoolLogin","Password")
        self.ObjSeleniumCommonLib.ScreenCapture("afterentervalue")
        self.ObjSeleniumCommonLib.Click("BBcoolLogin","btnLogin")
        self.ObjSeleniumCommonLib.ScreenCapture("afterlogin")
        a=self.ObjSeleniumCommonLib.GetTextMessage("BBcoolLogin","welcomelogintxt")
        import time
        time.sleep(10)
        print(str(a))


    #for mobile application
    def mobileappexample(self):
        try:
            print("****************************Vivek***************************")
            from appium import  webdriver
            desired_cap ={
                            "deviceName": "device",
                            "udid": "RZ8R31FPP5B",
                           "platformName": "Android",

                            "platformVersion": "11.0.0.0",
                            "appPackage": "secure.com.app.aasaan",
                            "appActivity": "secure.com.app.aasaan.coreengine.view.activities.SplashActivity",
                            "autoGrantPermissions": "true",
                            "autoAcceptAlerts": "true"
                            }
            #autoAcceptAlerts = true
            driver = webdriver.Remote("http://localhost:4723/wd/hub",desired_capabilities=desired_cap)
            import time
            time.sleep(10)
            #Locators
            # click on unaccepted
            unacceptedXpath = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout[2]/android.widget.RelativeLayout/android.widget.LinearLayout/android.widget.HorizontalScrollView/android.widget.LinearLayout/android.widget.FrameLayout[1]/android.widget.TextView"
            openTaskXpath = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout[2]/android.widget.RelativeLayout/android.widget.LinearLayout/android.widget.HorizontalScrollView/android.widget.LinearLayout/android.widget.FrameLayout[2]/android.widget.TextView"
            pendingTask = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.view.ViewGroup/android.widget.FrameLayout[2]/android.widget.RelativeLayout/android.widget.LinearLayout/androidx.viewpager.widget.ViewPager/android.widget.LinearLayout/android.widget.ListView/android.widget.LinearLayout[1]"
            startTask = "//*[@text = 'Start task']"
            step1Xpath = "//*[@text = 'Step1']"
            gtwaySerialNumber = "secure.com.app.aasaan:id/actGatewayBarCode"
            junctionBox = "secure.com.app.aasaan:id/rbJunctionBox"
            nextBtn = "secure.com.app.aasaan:id/btnSave"
            requestForRegistrationChkBox = "secure.com.app.aasaan:id/cb_gateway_mounted"
            requestForRegistrationSaveBtn = "secure.com.app.aasaan:id/btnReqRegistration"
            alreadySignedAlertAccept = "android:id/button1"
            msgAfterRequestforRegistration_Xpath = "/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/androidx.appcompat.widget.LinearLayoutCompat/android.widget.FrameLayout/android.widget.ScrollView/android.widget.LinearLayout/android.widget.TextView"
            acceptAlertForSaveBtn_id = "android:id/button1"
            registrationFailMsg_id = "secure.com.app.aasaan:id/tv_fail_message"
            registrationSuccessMSG_id = "secure.com.app.aasaan:id/tv_success_gateway_message"
            nextBtn_stepTwo_ID = "secure.com.app.aasaan:id/nextButton"
            discoveredMeter_Id = "secure.com.app.aasaan:id/tv_discovered_meter"
            acceptID = "secure.com.app.aasaan:id/btnAccept"

            #code to login
            driver.find_element_by_id("secure.com.app.aasaan:id/userName").send_keys("vivek")
            driver.find_element_by_id("secure.com.app.aasaan:id/password").send_keys("vivek")
            driver.find_element_by_id(("secure.com.app.aasaan:id/btnSignIn")).click()
            time.sleep(10)

            #Code to accept alert box for already signed in user
            try:
                WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.ID, alreadySignedAlertAccept)))
                driver.find_element_by_id(alreadySignedAlertAccept).click()
            except Exception as e:
                print(str(e))
                print("Accept alert not present.")
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)

            #Code to accept the task assigned from FFS
            WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, unacceptedXpath)))
            driver.find_element_by_xpath(unacceptedXpath).click()

            # Click on accept

            try:
                WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.ID, acceptID)))
                driver.find_element_by_id(acceptID).click()
            except Exception as e:
                print(str(e))
                print("Accept Button not present.")
                test_check.writeResultExclReport(self, actualData="Accept Button Not Present", dataToVerify="Accept button should be visible and clickable",
                                                 result="Fail")

                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)

            #Code to start task Step 1
            driver.find_element_by_xpath(openTaskXpath).click()
            WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, pendingTask)))

            driver.find_element_by_xpath(pendingTask).click()
            WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, startTask)))

            driver.find_element_by_xpath(startTask).click()
            WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.ID, gtwaySerialNumber)))

            driver.find_element_by_id(gtwaySerialNumber).send_keys("RND15088")
            driver.find_element_by_id(junctionBox).click()
            driver.find_element_by_id(nextBtn).click()



            #Read Message

            WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, msgAfterRequestforRegistration_Xpath)))
            #time.sleep(4)
            msgDisplayed = driver.find_element_by_xpath(msgAfterRequestforRegistration_Xpath).text
            driver.find_element_by_id(acceptAlertForSaveBtn_id).click()
            print("Message Displayed after clicking Request for Registration Button in step 1 = " +str(msgDisplayed))
            test_check.writeResultExclReport(self, actualData=msgDisplayed, dataToVerify="Saved successfully",
                                             result="Pass")

            time.sleep(4)
            driver.find_element_by_id(requestForRegistrationChkBox).click()
            driver.find_element_by_id(requestForRegistrationSaveBtn).click()


            #code to check if registration is passed
            try:
                WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.ID, registrationSuccessMSG_id)))
                successMessageDisplayed = driver.find_element_by_id(registrationSuccessMSG_id).text
                print("Registration message displayed is = " +str(successMessageDisplayed))
                test_check.writeResultExclReport(self, actualData=successMessageDisplayed, dataToVerify="Gateway number RNDXXXX was configured successfully.",
                                                 result="Pass")
                WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.ID, nextBtn_stepTwo_ID)))
                driver.find_element_by_id(nextBtn_stepTwo_ID).click()
                WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.ID, discoveredMeter_Id)))
                discoveredMeter = driver.find_element_by_id(discoveredMeter_Id).text
                print("Total Discovered meter is = " +str(discoveredMeter))
                discoveredMeter = discoveredMeter.split(" ")
                expectedMeterCount = self.FrameworkInitializer.GetTestDataFromDataTable("Actual Row Count")
                if discoveredMeter[0] == expectedMeterCount:
                    print("Test Case Passed, Data count Matched")
                    test_check.writeResultExclReport(self, actualData=expectedMeterCount,
                                                     dataToVerify=discoveredMeter[0],
                                                     result="Pass")
                else:
                    print("Test Case Failed, Data count MisMatched")
                    test_check.writeResultExclReport(self, actualData=expectedMeterCount,
                                                     dataToVerify=discoveredMeter[0],
                                                     result="Fail")



            except Exception as e:
                try:
                    WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.ID, registrationFailMsg_id)))
                    failMessage = driver.find_element_by_id(registrationFailMsg_id).text
                    print("Fail Message is = " + str(failMessage))
                    test_check.writeResultExclReport(self, actualData=failMessage,
                                                     dataToVerify="Gateway Registration Failed",
                                                     result="Fail")
                except Exception as e:
                    print(str(e))
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)


            #self.Appium_Android.ScreenCapture()
            #self.Appium_Android.SetValueInEditBox("vivek.purohit@securemeters.com", "BBcoolLoginMob", "Email")
            #self.Appium_Android.SetValueInEditBox("secure@123", "BBcoolLoginMob", "Password")
            #self.Appium_Android.ScreenCapture()
            #self.Appium_Android.Click("BBcoolLoginMob", "btnLogin")

        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        #self.Appium_Android.ScreenCapture()
        #self.FrameworkCommonLibrary.SelenuimWait(10)
        #from selenium import webdriver
        #drivers = webdriver.Chrome(executable_path="C:")
        #drivers.find_element_by_id().send_keys()
        #drivers.find_element_by_id().text
        #element = Select(ListBoxElement)
        #element.select_by_visible_text(strListItemValue)




    def UserManagement_VerifyRoles(self):
        try:
            #Login
            #Locators
            Tenant_id = "Tenant"
            UserName_id = "txtUserName"
            Password_id = "txtPwd"
            SubmitBtn_id = "// button[contains(text(), 'Login')]"
            manageUserLink_id = 'usermanagement'
            ImportGatewayAssetID = "gatewayassetimport"
            GatewayFirmwareUpgradeID = "upgradegatewayfirmware"
            AddScheduleXpath = "//span[contains(text(),'Add schedule')]"
            ManageUsersID = "usermanagement"
            ManageUserSubLinkXpath = "//body[1]/div[2]/section[1]/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[2]/ul[1]/li[2]/a[1]"
            DefineRoleXpath = "//span[contains(text(),'Define role')]"
            AddUserXpath = "//span[contains(text(),'Add user')]"
            DashboardXpath = "//a[contains(text(),'Dashboard')]"
            ConfigurationXpath = "//a[contains(text(),'Configuration')]"
            ReportsXpath = "//a[contains(text(),'Reports')]"
            logoutLinkBtnId = "userlink_btn"
            logoutlinkXpath = "//a[contains(text(),'Logout')]"
            pageTitleID = "pagetitle"
            temporaryPasswordXpath = "//*[@formcontrolname = 'oldPassword']"
            newPasswordXpath = "//*[@formcontrolname = 'password']"
            confirmPasswordXpath = "//*[@formcontrolname = 'confirmPassword']"
            saveBtnPasswordXpath = "//*[@type = 'submit' and @value = 'Save']"




            import configparser
            config = configparser.RawConfigParser()
            #config.read(r"C:\Users\44454\Desktop\Python Working Framework\eMBCBanswara\SeleniumAutomation\ObjectRepository\config.properties")
            self.FrameworkCommonLibrary.Driver.get("https://insightval.ewatch.online")
            WebDriverWait(self.FrameworkCommonLibrary.Driver, 100).until(EC.presence_of_element_located((By.ID, Tenant_id)))

            element = Select(self.FrameworkCommonLibrary.Driver.find_element_by_id(Tenant_id))
            element.select_by_visible_text(self.FrameworkInitializer.GetTestDataFromDataTable("Tenant"))

            self.FrameworkCommonLibrary.Driver.find_element_by_id(UserName_id).send_keys(self.FrameworkInitializer.GetTestDataFromDataTable("Email"))
            self.FrameworkCommonLibrary.Driver.find_element_by_id( Password_id).send_keys(self.FrameworkInitializer.GetTestDataFromDataTable("Password"))
            self.FrameworkCommonLibrary.Driver.find_element_by_xpath(SubmitBtn_id).click()

            # Code for wait
            WebDriverWait(self.FrameworkCommonLibrary.Driver, 10).until(EC.presence_of_element_located((By.ID, pageTitleID)))
            pageTitleText = self.FrameworkCommonLibrary.Driver.find_element_by_id(pageTitleID).text
            print(pageTitleText)
            LoginTitleDatabaseSheet = self.FrameworkInitializer.GetTestDataFromDataTable("Login Verify Message")
            LoginCredentials = str(self.FrameworkInitializer.GetTestDataFromDataTable("Email"))
            if pageTitleText.upper() == LoginTitleDatabaseSheet.upper():
                print("Login SuccessFully")
                test_check.writeResultExclReport(self, actualData= "Login User ID = " +LoginCredentials +" " +", Page Title = " +pageTitleText,dataToVerify=LoginTitleDatabaseSheet, result="Pass")

            else:
                print("Login Failed")
                test_check.writeResultExclReport(self, actualData=pageTitleText, dataToVerify=LoginTitleDatabaseSheet,result="Fail")

            # Change password
            userType = self.FrameworkInitializer.GetTestDataFromDataTable("User Type")
            if userType.upper() == "NEW":
                try:
                    tempPassword= self.FrameworkInitializer.GetTestDataFromDataTable("Temporary Password (In case of New User)")
                    newPassword = self.FrameworkInitializer.GetTestDataFromDataTable("New Password (In case of New User)")
                    WebDriverWait(self.FrameworkCommonLibrary.Driver, 10).until(EC.presence_of_element_located((By.XPATH, temporaryPasswordXpath)))
                    #Enter password
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(temporaryPasswordXpath).send_keys(tempPassword)
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(newPasswordXpath).send_keys(newPassword)
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(confirmPasswordXpath).send_keys(newPassword)
                    #Click save button
                    time.sleep(2)
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(saveBtnPasswordXpath).click()
                    time.sleep(4)
                    #Click on ok button
                    okBtnXpath = "//*[@class='active_btn' and @value = 'Ok']"
                    WebDriverWait(self.FrameworkCommonLibrary.Driver, 300).until(EC.presence_of_element_located((By.XPATH, okBtnXpath)))
                    messagePasswordChanged = self.FrameworkCommonLibrary.Driver.find_element_by_xpath("//*[@class='modal_bodytext']").text
                    print(messagePasswordChanged)
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(okBtnXpath).click()
                    time.sleep(4)
                    #Login again after changed password

                    element = Select(self.FrameworkCommonLibrary.Driver.find_element_by_id(Tenant_id))
                    element.select_by_visible_text(self.FrameworkInitializer.GetTestDataFromDataTable("Tenant"))
                    self.FrameworkCommonLibrary.Driver.find_element_by_id(UserName_id).send_keys(self.FrameworkInitializer.GetTestDataFromDataTable("Email"))
                    self.FrameworkCommonLibrary.Driver.find_element_by_id(Password_id).send_keys(self.FrameworkInitializer.GetTestDataFromDataTable("New Password (In case of New User)"))
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(SubmitBtn_id).click()

                    # Code for wait
                    WebDriverWait(self.FrameworkCommonLibrary.Driver, 10).until(EC.presence_of_element_located((By.ID, pageTitleID)))
                    pageTitleText = self.FrameworkCommonLibrary.Driver.find_element_by_id(pageTitleID).text
                    print(pageTitleText)
                    LoginTitleDatabaseSheet = self.FrameworkInitializer.GetTestDataFromDataTable("Login Verify Message")
                    LoginCredentials = str(self.FrameworkInitializer.GetTestDataFromDataTable("Email"))
                    if pageTitleText.upper() == LoginTitleDatabaseSheet.upper():
                        print("Login SuccessFully")
                        test_check.writeResultExclReport(self,
                                                         actualData="Login User ID = " + LoginCredentials + " " + ", Page Title = " + pageTitleText,
                                                         dataToVerify=LoginTitleDatabaseSheet, result="Pass")

                    else:
                        print("Login Failed")
                        test_check.writeResultExclReport(self, actualData=pageTitleText,
                                                         dataToVerify=LoginTitleDatabaseSheet, result="Fail")


                except Exception as e:
                    print(str(e))
                    print("Password not changed")
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)
            else:
                pass



            #Verify Login


            #Get Data from the database sheet
            ImportGatewayAsset = self.FrameworkInitializer.GetTestDataFromDataTable("Import Gateway Asset")
            GatewayFirmwareUpgrade = self.FrameworkInitializer.GetTestDataFromDataTable("Gateway Firmware Upgrade")
            AddSchedule = self.FrameworkInitializer.GetTestDataFromDataTable("Add Schedule")
            ManageUsers = self.FrameworkInitializer.GetTestDataFromDataTable("Manage Users")
            DefineRole = self.FrameworkInitializer.GetTestDataFromDataTable("Define Role")
            AddUser = self.FrameworkInitializer.GetTestDataFromDataTable("Add User")
            Dashboard = self.FrameworkInitializer.GetTestDataFromDataTable("Dashboard")
            Configuration = self.FrameworkInitializer.GetTestDataFromDataTable("Configuration")
            Reports = self.FrameworkInitializer.GetTestDataFromDataTable("Reports")

            #element = self.FrameworkCommonLibrary.Driver.find_element_by_id(ImportGatewayAssetID)
            #print("*************" +str(ImportGatewayAsset))
            #if ImportGatewayAsset.upper() == "YES":
                #try:
                    #if element is not None:
                        #print("[Info]: Import Gateway Link is present, Result = Pass")
                        #test_check.writeResultExclReport(self, actualData="Import Gateway Asset Link Present", dataToVerify=ImportGatewayAsset , result= "Pass")

                    #else:
                        #print("[Info]: Import Gateway Link is Not present, Result = Fail")
                        #test_check.writeResultExclReport(self, actualData="Import Gateway Asset Link Not Present", dataToVerify=ImportGatewayAsset , result= "Fail")
                #except Exception as e:
                    #print(str(e))
            #else:
                #print("[Info]: Import Gateway Link is Not present, Result = Pass")
                #test_check.writeResultExclReport(self, actualData="Import Gateway Asset Link is Not Present", dataToVerify=ImportGatewayAsset , result= "Pass")



            if GatewayFirmwareUpgrade.upper() == "YES":
                try:
                    element = self.FrameworkCommonLibrary.Driver.find_element_by_id(GatewayFirmwareUpgradeID)
                    if element is not None:
                        print("[Info]: GatewayFirmwareUpgrade Link is present, Result = Pass")
                        test_check.writeResultExclReport(self, actualData="GatewayFirmwareUpgrade Link Present",dataToVerify=GatewayFirmwareUpgrade, result="Pass")

                        #code for add schedule

                        if AddSchedule.upper() == "YES":
                            element = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(AddScheduleXpath)

                            if element is not None:
                                print("[Info]: AddSchedule Link is present, Result = Pass")
                                test_check.writeResultExclReport(self, actualData="AddSchedule Link Present",
                                                                 dataToVerify=AddSchedule, result="Pass")

                            else:
                                print("[Info]: AddSchedule is Not present, Result = Fail")
                                test_check.writeResultExclReport(self, actualData="AddSchedule Not Present",
                                                                 dataToVerify=AddSchedule, result="Fail")
                        else:
                            print("[Info]: AddSchedule Link is Not present, Result = Pass")
                            test_check.writeResultExclReport(self, actualData="AddSchedule Link is Not Present",
                                                             dataToVerify=AddSchedule, result="Pass")


                    else:
                        print("[Info]: GatewayFirmwareUpgrade is Not present, Result = Fail")
                        test_check.writeResultExclReport(self, actualData="GatewayFirmwareUpgrade Not Present",dataToVerify=GatewayFirmwareUpgrade, result="Fail")
                except Exception as e:
                    print(str(e))
                    test_check.writeResultExclReport(self, actualData="Exception Caught while test execution",
                                                     dataToVerify="Validating for Firmware Upgrade and sublinks", result="Fail")
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)
            else:
                print("[Info]: GatewayFirmwareUpgrade Link is Not present, Result = Pass")
                test_check.writeResultExclReport(self, actualData="GatewayFirmwareUpgrade Link is Not Present", dataToVerify=GatewayFirmwareUpgrade, result="Pass")




            #Code for Manage Users

            if ManageUsers.upper() == "YES":
                try:

                    if element is not None:
                        element = self.FrameworkCommonLibrary.Driver.find_element_by_id(ManageUsersID)
                        print("[Info]: ManageUsers Link is present, Result = Pass")
                        test_check.writeResultExclReport(self, actualData="ManageUsers Link Present",
                                                         dataToVerify=ManageUsers, result="Pass")

                        #Code for Define Role
                        self.FrameworkCommonLibrary.Driver.find_element_by_id(ManageUsersID).click()
                        time.sleep(2)

                        if DefineRole.upper() == "YES":
                            element = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(DefineRoleXpath)
                            if element is not None:
                                print("[Info]: DefineRole Link is present, Result = Pass")
                                test_check.writeResultExclReport(self, actualData="DefineRole Link Present",
                                                                 dataToVerify=DefineRole, result="Pass")
                            else:
                                print("[Info]: DefineRole is Not present, Result = Fail")
                                test_check.writeResultExclReport(self, actualData="DefineRole Not Present",
                                                                 dataToVerify=DefineRole, result="Fail")
                        else:
                            print("[Info]: DefineRole Link is Not present, Result = Pass")
                            test_check.writeResultExclReport(self, actualData="AddSchedule Link is Not Present",
                                                             dataToVerify=DefineRole, result="Pass")

                        #code for Add user


                        if AddUser.upper() == "YES":
                            self.FrameworkCommonLibrary.Driver.find_element_by_xpath(ManageUserSubLinkXpath).click()
                            time.sleep(3)
                            element = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(AddUserXpath)

                            if element is not None:
                                print("[Info]: AddUser Link is present, Result = Pass")
                                test_check.writeResultExclReport(self, actualData="AddUser Link Present",
                                                                     dataToVerify=AddUser, result="Pass")
                            else:
                                print("[Info]: AddUser is Not present, Result = Fail")
                                test_check.writeResultExclReport(self, actualData="AddUser Not Present",
                                                                     dataToVerify=AddUser, result="Fail")
                        else:
                            print("[Info]: AddUser Link is Not present, Result = Pass")
                            test_check.writeResultExclReport(self, actualData="AddUser Link is Not Present",
                                                                 dataToVerify=AddUser, result="Pass")


                    else:
                        print("[Info]: ManageUsers is Not present, Result = Fail")
                        test_check.writeResultExclReport(self, actualData="ManageUsers Not Present",
                                                         dataToVerify=ManageUsers, result="Fail")
                except Exception as e:
                    print(str(e))
                    print(str(e))
                    test_check.writeResultExclReport(self, actualData="Exception Caught while test execution",
                                                     dataToVerify="Validating for ManageUsers and sublinks",
                                                     result="Fail")
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)


            else:
                print("[Info]: ManageUsers Link is Not present, Result = Pass")
                test_check.writeResultExclReport(self, actualData="ManageUsers Link is Not Present",
                                                 dataToVerify=ManageUsers, result="Pass")

            #Code fo dashboard

            if Dashboard.upper() == "YES":
                element = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(DashboardXpath)
                try:
                    if element is not None:
                        print("[Info]: Dashboard Link is present, Result = Pass")
                        test_check.writeResultExclReport(self, actualData="Dashboard Link Present",
                                                         dataToVerify=Dashboard, result="Pass")

                    else:
                        print("[Info]: Dashboard is Not present, Result = Fail")
                        test_check.writeResultExclReport(self, actualData="Dashboard Not Present",
                                                         dataToVerify=Dashboard, result="Fail")
                except Exception as e:
                    print(str(e))
                    print(str(e))
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)
            else:
                print("[Info]: Dashboard Link is Not present, Result = Pass")
                test_check.writeResultExclReport(self, actualData="Dashboard Link is Not Present",
                                                 dataToVerify=Dashboard, result="Pass")

            #Code for Configuration

            if Configuration.upper() == "YES":
                element = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(ConfigurationXpath)
                try:
                    if element is not None:
                        print("[Info]: Configuration Link is present, Result = Pass")
                        test_check.writeResultExclReport(self, actualData="Configuration Link Present",
                                                         dataToVerify=Configuration, result="Pass")

                    else:
                        print("[Info]: Configuration is Not present, Result = Fail")
                        test_check.writeResultExclReport(self, actualData="Configuration Not Present",
                                                         dataToVerify=Configuration, result="Fail")
                except Exception as e:
                    print(str(e))
                    print(str(e))
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)
            else:
                print("[Info]: Configuration Link is Not present, Result = Pass")
                test_check.writeResultExclReport(self, actualData="Configuration Link is Not Present",
                                                 dataToVerify=Configuration, result="Pass")


            if Reports.upper() == "YES":
                try:
                    element = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(ReportsXpath)
                    if element is not None:
                        print("[Info]: Reports Link is present, Result = Pass")
                        test_check.writeResultExclReport(self, actualData="Reports Link Present",
                                                         dataToVerify=Reports, result="Pass")

                    else:
                        print("[Info]: Reports is Not present, Result = Fail")
                        test_check.writeResultExclReport(self, actualData="Reports Not Present",
                                                         dataToVerify=Reports, result="Fail")
                except Exception as e:
                    print(str(e))
                    test_check.writeResultExclReport(self, actualData="Result Fail",
                                                     dataToVerify=Reports, result="Fail")
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)
            else:
                print("[Info]: Reports Link is Not present, Result = Pass")
                test_check.writeResultExclReport(self, actualData="Reports Link is Not Present",
                                                 dataToVerify=Reports, result="Pass")

            self.FrameworkCommonLibrary.Driver.find_element_by_id(logoutLinkBtnId).click()
            time.sleep(2)
            self.FrameworkCommonLibrary.Driver.find_element_by_xpath(logoutlinkXpath).click()

            #try:
                #WebDriverWait(self.FrameworkCommonLibrary.Driver, 100).until(EC.presence_of_element_located((By.ID, manageUserLink_id)))
                #self.FrameworkCommonLibrary.Driver.find_element_by_id(manageUserLink_id).click()
                #element = self.FrameworkCommonLibrary.Driver.find_element_by_id(manageUserLink_id)
                #if element is not None:
                    #print("Element Present, and Verified")
            #except Exception as e:
                #print(str(e))
                #exc_type, exc_obj, exc_tb = sys.exc_info()
                #fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                #print(exc_type, fname, exc_tb.tb_lineno)
        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

    # Code for database value comparison
    def gatewayAssetImport_DatabaseComparison(self):
        try:
            gtwySerialNo = self.FrameworkInitializer.GetTestDataFromDataTable("Gateway Serial No")
            import pyodbc
            conn = pyodbc.connect(
                'DRIVER={PostgreSQL Unicode};SERVER=10.10.103.250;Port=5432;DATABASE=FFS_Test;UID=postgres;PWD=postgres')
            cursor = conn.cursor()
            sqlQuery1 = "SELECT " + '"nProjectid",' + '"gwSerialNumber",' + '"gwMacAddress",' + '"hwType",' + '"productName",' + " createddatetime FROM public.gatewayforissue " + "where " + '"gwSerialNumber"' + "=" + "'" + gtwySerialNo + "'" + " order By createddatetime DESC;"
            print("[Info]: SQl Query is = " + sqlQuery1)
            cursor.execute(sqlQuery1)
            rowcount = cursor.rowcount
            print("[Info]: Row Count is = " + str(rowcount))
            if rowcount == 0:
                print("[Info]: No result fetched in GatewayForIssue Table")
            else:
                counter = 0
                for row in cursor:
                    counter = counter + 1
                    print("[Info]: Result = Pass, Rows Imported from GatewayForIssue Table = : ")
                    print("[Details: Info]: Database row imported = " + str(row))
                    print("[Details: Info]: Gateway Serial No. is = " + str(row[1]))
                    print("[Details: Info]: Created Date Time is = " + str(row[5]))
                    print("[Details: Info]: Result = Pass")

            # Code for gateway keys mst
            conn1 = pyodbc.connect('DRIVER={PostgreSQL Unicode};SERVER=10.10.103.151;Port=1606;DATABASE=postgres;UID=postgres;PWD=postgres')
            cursor = conn1.cursor()
            sqlQuery1 = "SELECT gatewayid, importdate, gatewaymacid, gatewayserialnumber, gatewayexternalserialnumber FROM smarthomes.tb_gatewaykeysmst where gatewayexternalserialnumber = " +"'" +gtwySerialNo + "'"+" order by importdate DESC;"
            print("[Info]: SQl Query is = " + sqlQuery1)
            cursor.execute(sqlQuery1)
            rowcount = cursor.rowcount
            print("[Info]: Row Count is = " + str(rowcount))
            if rowcount == 0:
                print("[Info]: No result fetched in GatewayKeysMst table")
            else:
                counter = 0
                for row in cursor:
                    counter = counter + 1
                    print("[Info]: Result = Pass, Rows Imported from GatewayKeysMst table = : ")
                    print("[Details: Info]: Database row imported = " + str(row))
                    print("[Details: Info]: Gateway Serial No. is = " + str(row[4]))
                    print("[Details: Info]: Imported Date Time is = " + str(row[1]))
                    print("[Details: Info]: Result = Pass")


            # Code for gatewaymst
            conn2 = pyodbc.connect('DRIVER={PostgreSQL Unicode};SERVER=10.10.103.151;Port=1606;DATABASE=postgres;UID=postgres;PWD=postgres')
            cursor = conn2.cursor()
            sqlQuery1 = "SELECT gatewayid, gatewayexternalserialnumber, gatewaymacid, importdate FROM smarthomes.tb_gatewaymst where gatewayexternalserialnumber = " +"'" +gtwySerialNo + "'"+" order by importdate DESC;"
            print("[Info]: SQl Query is = " + sqlQuery1)
            cursor.execute(sqlQuery1)
            rowcount = cursor.rowcount
            print("[Info]: Row Count is = " + str(rowcount))
            if rowcount == 0:
                self.ObjTestResult.WriteResultInResultReport("No result fetched", TestCaseResults.Fail)
            else:
                counter = 0
                for row in cursor:
                    counter = counter + 1
                    print("[Info]: Result = Pass, Rows imported from GatewayMst table")
                    print("[Details: Info]: Database row imported = " + str(row))
                    print("[Details: Info]: Gateway Serial No. is = " + str(row[1]))
                    print("[Details: Info]: Imported Date Time is = " + str(row[3]))
                    print("[Details: Info]: Result = Pass")
                    self.ObjTestResult.WriteResultInResultReport("Result Fetched, Data Matched", TestCaseResults.Pass)

        except Exception as e:
            #self.ScreenCapture("Fail")
            self.ObjTestResult.WriteResultInResultReport("Exception in control object:" + str(e), TestCaseResults.Fail)
            print(str(e))

    def CreateExcelReport(self):
        print("***************************************inside create excel report")
        try:
            import pathlib
            ResultFolderPath = None
            if pathlib.Path("..\\..\\..\\TestResultReports").exists():
                print("[Info]: TestResultReports Folder exist")
            else:
                pathlib.Path("..\\..\\..\\TestResultReports").mkdir()
                print("[Success]: TestResultReports Folder Created")
        except Exception as e:
            print(str(e))

        #Create folder for Final Report Folder
        try:
            import pathlib
            ResultFolderPath = None
            if pathlib.Path("..\\..\\..\\TestResultReports\\Final Report Folder").exists():
                print("[Info]: Final Report Folder Folder exist")
            else:
                pathlib.Path("..\\..\\..\\TestResultReports\\Final Report Folder").mkdir()
                print("[Success]: Final Report Folder Folder Created")
        except Exception as e:
            print(str(e))

        from datetime import datetime
        # datetime object containing current date and time
        now = datetime.now()
        print("[Info]: Checking and Creating required directories.")
        print("[Info]: Current Date Time =", now)
        dt_string = now.strftime("%d_%m_%Y_%H_%M_%S")
        print("[Info]: Current Date Time in required format = ", dt_string)

        import xlsxwriter
        #self.ObjSeleniumConfiguration.workbook = Workbook()
        self.ObjSeleniumConfiguration.workbookReport = Workbook()
        self.ObjSeleniumConfiguration.worksheetReport = worksheet
        self.ObjSeleniumConfiguration.workbookReport = xlsxwriter.Workbook("..\\..\\..\\TestResultReports\\Final Report Folder\\Final_Report" + dt_string + ".xlsx")
        self.ObjSeleniumConfiguration.Excel_File_Name = "..\\..\\..\\TestResultReports\\Final Report Folder\\Final_Report" + dt_string + ".xlsx"
        self.ObjSeleniumConfiguration.worksheetReport = self.ObjSeleniumConfiguration.workbookReport.add_worksheet()
        print("[Info]: Work Sheet Created")
        # Adding Font Size and Cell Format to the excel created
        self.ObjSeleniumConfiguration.cell_format = self.ObjSeleniumConfiguration.workbookReport.add_format({'bold': True, 'font_color': '#000000', 'text_wrap': True, 'bg_color': '#D3D3D3', 'border': True})
        self.ObjSeleniumConfiguration.worksheetReport.set_column(0, 10, 30)
        self.ObjSeleniumConfiguration.cell_format_1 = self.ObjSeleniumConfiguration.workbookReport.add_format({'text_wrap': True, 'border': True})
        self.ObjSeleniumConfiguration.cell_format_Pass = self.ObjSeleniumConfiguration.workbookReport.add_format({'bold': True, 'text_wrap': True, 'border': True, 'bg_color': '#008000'})
        self.ObjSeleniumConfiguration.cell_format_Fail = self.ObjSeleniumConfiguration.workbookReport.add_format({'bold': True, 'text_wrap': True, 'border': True, 'bg_color': '#FF0000'})
        # self.frameworkInitializer.worksheet.conditional_format('H1:H100', {'type': 'text','criteria': 'equal to','value': 'Pass', 'format': self.frameworkInitializer.cell_format_Pass})

        # Setting the Header row of the Excel
        self.ObjSeleniumConfiguration.row = 0
        self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 1, "Test Scenario",self.ObjSeleniumConfiguration.cell_format)
        self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 2, "Result",self.ObjSeleniumConfiguration.cell_format)
        self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 3, "Expected Data",self.ObjSeleniumConfiguration.cell_format)
        self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 4, "Actual Data", self.ObjSeleniumConfiguration.cell_format)
        self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 0, str("Test Data ID"),self.ObjSeleniumConfiguration.cell_format)
        self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 5, "Execution Time", self.ObjSeleniumConfiguration.cell_format)
        self.ObjSeleniumConfiguration.row = self.ObjSeleniumConfiguration.row + 1

    def writeResultExclReport(self, actualData, dataToVerify, result):
        scenario = self.ObjSeleniumConfiguration.TestScenarioName
        self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 1, scenario,self.ObjSeleniumConfiguration.cell_format_1)
        testDataID = self.ObjSeleniumConfiguration.TestCaseReferenceNo
        #testDataID = self.FrameworkInitializer.GetTestDataFromDataTable("Test Data ID")
        #self.ObjSeleniumConfiguration.worksheet.write_url(self.ObjSeleniumConfiguration.row, 0, testDataID,self.ObjSeleniumConfiguration.cell_format_1)
        location = r'external:C:\Users\44454\Desktop\Python Working Framework\embc\Insight\SeleniumAutomation\DataBase' +"\\" + self.ObjSeleniumConfiguration.TestDataExcelFile
        self.ObjSeleniumConfiguration.worksheetReport.write_url(self.ObjSeleniumConfiguration.row, 0, location,string=testDataID)
        #self.ObjSeleniumConfiguration.worksheet.write_url(self.ObjSeleniumConfiguration.row, 0, r'external:C:\Users\44454\Desktop\Python Working Framework\embc\Insight\SeleniumAutomation\ControlFile\NewOR\WebTestCases.xlsx', string=testDataID)
        if result == "Pass":
            self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 2, "Pass", self.ObjSeleniumConfiguration.cell_format_Pass)
        elif result == "Fail":
            self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 2, "Fail", self.ObjSeleniumConfiguration.cell_format_Fail)


        #Message = self.FrameworkInitializer.GetTestDataFromDataTable("Message")
        self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 3, dataToVerify,self.ObjSeleniumConfiguration.cell_format_1)
        #actualData = "Hard Data ------ Schedule Saved Successfully."
        self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 4, actualData,self.ObjSeleniumConfiguration.cell_format_1)
        from datetime import datetime
        # datetime object containing current date and time
        now = datetime.now()
        dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
        self.ObjSeleniumConfiguration.worksheetReport.write(self.ObjSeleniumConfiguration.row, 5, dt_string,self.ObjSeleniumConfiguration.cell_format_1)
        self.ObjSeleniumConfiguration.row = self.ObjSeleniumConfiguration.row + 1
        #self.ObjSeleniumConfiguration.workbookReport.close()


    def testingExec(self):
        test_check.writeResultExclReport(self,"Testing to print actual data", result= "Fail")

    def UserManagement_AddRoles(self):

        try:
            import time

            #Define Locators
            configurationlink_id = "ConfigId"
            manageUserLink_id = "usermanagement"
            defineRoleLink_xpath = "//span[contains(text(),'Define role')]"
            roleNameEdtBox_xpath = "//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/form[1]/div[2]/div[1]/div[1]/input[1]"
            descriptionEdtBox_xpath="//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/form[1]/div[2]/div[1]/div[2]/input[1]"
            #Locators for checkbox
            configuration= "//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[4]/div[1]/div[1]/div[2]/div[2]/div[1]/div[2]/div[1]/label[1]/span[1]"
            gatewayAssetImport_add = "//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[4]/div[2]/div[2]/div[3]/feature-list[1]/div[1]/div[3]/div[2]/label[1]/span[1]"
            gatewayFirmwareUpgrade_add = "//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[4]/div[2]/div[2]/div[3]/feature-list[1]/div[1]/div[3]/div[2]/label[1]/span[1]"
            gatewayFirmwareUpgrade_view ="//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[4]/div[2]/div[2]/div[3]/feature-list[1]/div[1]/div[3]/div[1]/label[1]/span[1]"
            userManagement_add = "//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[4]/div[2]/div[2]/div[3]/feature-list[1]/div[2]/div[3]/div[2]/label[1]/span[1]"
            userManagement_view = "//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[4]/div[2]/div[2]/div[3]/feature-list[1]/div[2]/div[3]/div[1]/label[1]/span[1]"
            dashboard_xpath = "//span[contains(text(),'Dashboard')]"
            report_xpath = "//span[contains(text(),'Report')]"
            rebootGtwy_xpath = "//span[contains(text(),'Reboot gateway')]"
            DwnldGtwyLogFile = "//span[contains(text(),'Download gateway log file')]"
            reportChkBox = "//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[4]/div[2]/div[2]/div[3]/feature-list[1]/div[1]/div[3]/div[1]/label[1]/span[1]"
            searchInputXpath = "//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[4]/div[1]/div[1]/input[1]"
            searchBtnXpath = "//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[4]/div[1]/div[1]/input[2]"
            searchResultXpath = "//body/div[@id='mainarea']/section[@id='formareamain']/div[1]/div[1]/div[3]/div[1]/section[1]/app-root[1]/users[1]/ng-component[1]/div[5]/jqxgrid[1]/div[1]/div[1]/div[1]/div[4]/div[2]/div[1]/div[1]/div[4]/div[1]/a[1]/img[1]"
            editLinkXpath = "//a[contains(text(),'Edit')]"
            #test


            saveBtn_xpath = "//*[@class='active_btn' and @value='Save']"
            verifyMsg_xpath = "//*[@class='modal_bodytext']"
            okBtn_xpath = "//*[@class='active_btn' and @value='Done']"
            errorMessageXpath = "//*[@class='alert alert-danger']"
            errorModuleMessageXpath = "//div[contains(text(),'Select at least one module access right')]"

            #self.FrameworkCommonLibrary.Driver.get("https://www.google.com")
            # Clicked on Configuration Link
            self.FrameworkCommonLibrary.Driver.find_element_by_id(configurationlink_id).click()
            time.sleep(4)
            # Clicked on Manage User Link
            self.FrameworkCommonLibrary.Driver.find_element_by_id(manageUserLink_id).click()
            time.sleep(4)
            operationType = self.FrameworkInitializer.GetTestDataFromDataTable("Operation")
            if operationType.upper() == "EDIT":
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(searchInputXpath).send_keys(self.FrameworkInitializer.GetTestDataFromDataTable("Role Name (if Search for edit)"))
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(searchBtnXpath).click()
                time.sleep(2)
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(searchResultXpath).click()
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(editLinkXpath).click()
                time.sleep(4)


            elif operationType.upper() == "ADD":
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(defineRoleLink_xpath).click()

            #Clicked on Define Roles

            time.sleep(4)
            # Enter the Role name and Description
            if self.FrameworkInitializer.GetTestDataFromDataTable("Role_Name") is None:
                pass
            else:
              self.FrameworkCommonLibrary.Driver.find_element_by_xpath(roleNameEdtBox_xpath).clear()
              roleName = str(self.FrameworkInitializer.GetTestDataFromDataTable("Role_Name"))
              self.FrameworkCommonLibrary.Driver.find_element_by_xpath(roleNameEdtBox_xpath).send_keys(roleName)

            if self.FrameworkInitializer.GetTestDataFromDataTable("Description") is None:
                pass
            else:
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(descriptionEdtBox_xpath).clear()
                description = str(self.FrameworkInitializer.GetTestDataFromDataTable("Description"))
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(descriptionEdtBox_xpath).send_keys(description)


            time.sleep(3)
            # Read the test data
            checkBox = self.FrameworkInitializer.GetTestDataFromDataTable("XpathSequencing")
            checkBoxLst = checkBox.split(",")
            print(checkBoxLst)
            for i in range(len(checkBoxLst)):
                checkBox_Locator = self.FrameworkCommonLibrary.Locators("UserManagementAddRoles","AddRoles",checkBoxLst[i])
                print(checkBox_Locator)
                self.frameworkInitializer.Driver.find_element_by_xpath(checkBox_Locator[1]).click()
                time.sleep(1)
            time.sleep(2)
            self.FrameworkCommonLibrary.Driver.find_element_by_xpath(saveBtn_xpath).click()    
            """#Get the test data for the checkbox from the excel
            gtwyAsstImport = self.FrameworkInitializer.GetTestDataFromDataTable("GatewayAssetImport").split(",")
            gtwyFrmwareUpgrade = self.FrameworkInitializer.GetTestDataFromDataTable("GatewayFirmwareUpgrade").split(",")
            usrMgmt = self.FrameworkInitializer.GetTestDataFromDataTable("UserManagement").split(",")
            rebootGateway = self.FrameworkInitializer.GetTestDataFromDataTable("Reboot Gateway")
            downloadGatewayLogFile = self.FrameworkInitializer.GetTestDataFromDataTable("Download Gateway Log File")
            report = self.FrameworkInitializer.GetTestDataFromDataTable("Report")

            print("Gateway Asset import is = " +str(gtwyAsstImport))
            print("Gateway Firmware Upgrade is = " +str(gtwyFrmwareUpgrade))
            print("User management is = " +str(usrMgmt))

            #Check condition for checkbox print
            if len(gtwyAsstImport) == 1:
                if gtwyAsstImport[0].upper() == "ADD":
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(gatewayAssetImport_add).click()
                else:
                    pass

            else:
                pass
            if len(gtwyFrmwareUpgrade) == 1:
                if gtwyFrmwareUpgrade[0].upper() == "ADD":
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(gatewayFirmwareUpgrade_add).click()
                elif gtwyFrmwareUpgrade[0].upper() == "VIEW":
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(gatewayFirmwareUpgrade_view).click()
                else:
                    pass


            else:
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(gatewayFirmwareUpgrade_add).click()
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(gatewayFirmwareUpgrade_view).click()

            if len(usrMgmt) == 1:
                if usrMgmt[0].upper() == "ADD":
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(userManagement_add).click()
                elif usrMgmt[0].upper() == "VIEW":
                    self.FrameworkCommonLibrary.Driver.find_element_by_xpath(userManagement_view).click()
                else:
                    pass
            else:
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(userManagement_add).click()
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(userManagement_view).click()

            self.FrameworkCommonLibrary.Driver.find_element_by_xpath(dashboard_xpath).click()

            if rebootGateway.upper() == "ADD":

                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(rebootGtwy_xpath).click()
            else:
                pass


            if downloadGatewayLogFile.upper() == "ADD":
                self.FrameworkCommonLibrary.Driver.find_element_by_xpath(DwnldGtwyLogFile).click()
            else:
                pass
            #self.FrameworkCommonLibrary.Driver.find_element_by_xpath(report_xpath).click()
            #if report.upper() == "VIEW":
            #   self.FrameworkCommonLibrary.Driver.find_element_by_xpath(reportChkBox).click()
            #else:
            #    pass

            # Click on save button
            time.sleep(2)
            self.FrameworkCommonLibrary.Driver.find_element_by_xpath(saveBtn_xpath).click()
            time.sleep(4)
            testText = self.FrameworkInitializer.GetTestDataFromDataTable("Role_Name")
            if len(roleName) == 0:
                print("Inside Role Name blank")
                actualMsg = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(errorMessageXpath).text
                expectedMsg = self.FrameworkInitializer.GetTestDataFromDataTable("Verify_Message")
                print("Expected Message = " + str(expectedMsg))
                print("Actual Message = " + str(actualMsg))
                if actualMsg == expectedMsg:
                    print("Text Matched, Result = Pass")
                    test_check.writeResultExclReport(self, actualData=actualMsg, dataToVerify=expectedMsg,
                                                     result="Pass")

                else:
                    print("Text Matched, Result = Fail")
                    test_check.writeResultExclReport(self, actualData=actualMsg, dataToVerify=expectedMsg,result="Fail")

            elif len(description) == 0:
                actualMsg = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(errorMessageXpath).text
                expectedMsg = self.FrameworkInitializer.GetTestDataFromDataTable("Verify_Message")
                print("Expected Message = " + str(expectedMsg))
                print("Actual Message = " + str(actualMsg))
                if actualMsg == expectedMsg:
                    print("Text Matched, Result = Pass")
                    test_check.writeResultExclReport(self, actualData=actualMsg, dataToVerify=expectedMsg,
                                                     result="Pass")

                else:
                    print("Text Matched, Result = Fail")
                    test_check.writeResultExclReport(self, actualData=actualMsg, dataToVerify=expectedMsg,
                                                     result="Fail")
            elif len(self.FrameworkInitializer.GetTestDataFromDataTable("GatewayAssetImport")) == 0:

                actualMsg = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(errorModuleMessageXpath).text
                expectedMsg = self.FrameworkInitializer.GetTestDataFromDataTable("Verify_Message")
                print("Expected Message = " + str(expectedMsg))
                print("Actual Message = " + str(actualMsg))
                if actualMsg == expectedMsg:
                    print("Text Matched, Result = Pass")
                    test_check.writeResultExclReport(self, actualData=actualMsg, dataToVerify=expectedMsg,result="Pass")

                else:
                    print("Text Matched, Result = Fail")
                    test_check.writeResultExclReport(self, actualData=actualMsg, dataToVerify=expectedMsg,
                                                     result="Fail")

            else:
                actualMsg = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(verifyMsg_xpath).text
                expectedMsg = self.FrameworkInitializer.GetTestDataFromDataTable("Verify_Message")
                print("Expected Message = " + str(expectedMsg))
                print("Actual Message = " + str(actualMsg))
                if actualMsg == expectedMsg:
                    print("Text Matched, Result = Pass")
                    test_check.writeResultExclReport(self, actualData=actualMsg, dataToVerify=expectedMsg,
                                                     result="Pass")

                else:
                    print("Text Matched, Result = Fail")
                    test_check.writeResultExclReport(self, actualData=actualMsg, dataToVerify=expectedMsg,
                                                     result="Fail")"""





            #Click on Ok button
            WebDriverWait(self.FrameworkCommonLibrary.Driver, 10).until(EC.presence_of_element_located((By.XPATH, okBtn_xpath)))
            self.FrameworkCommonLibrary.Driver.find_element_by_xpath(okBtn_xpath).click()
            print("******************** Clicked on Done Button")

            #Code to verify in view table:


            #Select the checkbox
            #self.FrameworkCommonLibrary.Driver.find_element_by_xpath(configuration).click()

        except Exception as e:
            #self.ScreenCapture("Fail")
            self.ObjTestResult.WriteResultInResultReport(
                "Exception in control object:" + str(e), TestCaseResults.Fail)
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            #WebDriverWait(self.FrameworkCommonLibrary.Driver, 10).until(EC.presence_of_element_located((By.XPATH, okBtn_xpath)))
            #self.FrameworkCommonLibrary.Driver.find_element_by_xpath(okBtn_xpath).click()



    def sqlQueryExecution(self):
        try:
            #Write to the database sheet
            condition_dataFetchToNextRow = self.FrameworkInitializer.GetTestDataFromDataTable("Fetch data to next Row")
            path = self.ObjSeleniumConfiguration.TestDataFolderPath + "\\" +"Insight_SQLquery.xlsx"
            #print(path)
            from openpyxl import load_workbook

            # load excel file
            workbook = load_workbook(filename=path)

            # open workbook
            sheet = workbook.active
            connectionString = self.FrameworkInitializer.GetTestDataFromDataTable("Connection String")
            sqlQuery= self.FrameworkInitializer.GetTestDataFromDataTable("SQL Query")
            substring = "XXX"
            fullstring = str(sqlQuery)
            replaceStringDetails = str(self.FrameworkInitializer.GetTestDataFromDataTable("Condition"))
            if substring in fullstring:
                print("Found!")
                sqlQuery = fullstring.replace("XXX", replaceStringDetails)
                print(fullstring.replace("XXX", replaceStringDetails))


            expectedRowCount = str(self.FrameworkInitializer.GetTestDataFromDataTable("Expected Row Count"))
            dataColumnMatch = self.FrameworkInitializer.GetTestDataFromDataTable("Data Match")
            expectedData = self.FrameworkInitializer.GetTestDataFromDataTable("Expected Data")
            import pyodbc
            conn = pyodbc.connect(connectionString)
            cursor = conn.cursor()
            print("sqlQuery is = " +sqlQuery)
            cursor.execute(sqlQuery)
            conn.commit()
            rowcount = cursor.rowcount
            print("Expected Row count is = " +expectedRowCount )
            print("[Info]: Row Count is = " + str(rowcount))


            #***************
            #print("Current Iteration is = " +str(self.ObjSeleniumConfiguration.intCurrentIteration))
            self.ObjSeleniumConfiguration.intCurrentIteration = self.ObjSeleniumConfiguration.intCurrentIteration +1
            counter = 0
            rowDataFetched = []
            try:
                for row in cursor:
                    counter = counter + 1
                    print("[Info]: Database row imported = " + str(row))
                    rowDataFetched.append(str(row))
            except Exception as e:
                print(str(e))
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)

            print(rowDataFetched)

            #write row count
            sheetColumnName = "K" + str(self.ObjSeleniumConfiguration.intCurrentIteration)
            sheet[sheetColumnName] = rowcount


            #write the fetched data

            if condition_dataFetchToNextRow.upper() == "YES":
                self.ObjSeleniumConfiguration.intCurrentIteration = self.ObjSeleniumConfiguration.intCurrentIteration + 1
                sheetColumnName_DataFetched = "G" + str(self.ObjSeleniumConfiguration.intCurrentIteration)
                #sheet[sheetColumnName_DataFetched] = rowDataFetched
                print(len(rowDataFetched))
                strippedString = ''.join(rowDataFetched)
                strippedString = strippedString.replace("(", "").replace(")", "").replace(" ", "")[0:-1]
                print(strippedString)
                sheet[sheetColumnName_DataFetched] = strippedString
            #for z in range(len(rowDataFetched)):
                #sheet[sheetColumnName_DataFetched] = str(rowDataFetched)[1:-1]
            else:
                sheetColumnName_DataFetched = "L" + str(self.ObjSeleniumConfiguration.intCurrentIteration)
                # sheet[sheetColumnName_DataFetched] = rowDataFetched
                print(len(rowDataFetched))
                strippedString = ''.join(rowDataFetched)
                strippedString = strippedString.replace("(", "").replace(")", "").replace(" ", "")[0:-1]
                print(strippedString)
                sheet[sheetColumnName_DataFetched] = strippedString

            # save the file
            workbook.save(filename=path)
            #***************

            try:
                if int(rowcount) == int(expectedRowCount):
                    print("Row Count matched, Result = Pass")
                    test_check.writeResultExclReport(self, actualData=int(rowcount),
                                                     dataToVerify=int(expectedRowCount),
                                                     result="Pass")
                else:
                    test_check.writeResultExclReport(self, actualData=int(rowcount),
                                                     dataToVerify=int(expectedRowCount),
                                                     result="Fail")
            except Exception as e:
                print(str(e))
            if rowcount == 0:
                print("No Result Fetched")
            else:
                counter = 0
                for row in cursor:
                    counter = counter + 1
                    print("Data column match value = " + str(dataColumnMatch))
                    print("[Details: Info]: Gateway Serial No. is = " + str(row[dataColumnMatch]))
                    actualDatafromQuery = str(row[dataColumnMatch])
                    if int(actualDatafromQuery) == int(expectedData):
                        print("Final Result = pass")
                        test_check.writeResultExclReport(self, actualData=int(actualDatafromQuery), dataToVerify=int(expectedData),
                                                         result="Pass")
                    else:
                        print("Final result = Fail")
                        test_check.writeResultExclReport(self, actualData=int(actualDatafromQuery),
                                                         dataToVerify=int(expectedData),
                                                         result="Fail")

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

    def switchingtobrowser(self):
        self.FrameworkCommonLibrary.SelenuimWait(10)
        self.ObjSeleniumConfiguration.BrowserName = "Chrome"
        optionsForGoogleChrome = webdriver.ChromeOptions()
        optionsForGoogleChrome.add_argument("ignore-certificate-errors")
        optionsForGoogleChrome.add_argument("--start-maximized")
        optionsForGoogleChrome.add_argument("--no-sandbox")
        optionsForGoogleChrome.add_argument("--disable-dev-shm-usage")
        Drivers = webdriver.Chrome(
            executable_path=self.ObjSeleniumConfiguration.SeleniumLibraryFolderPath + "/chromedriver.exe",
            chrome_options=optionsForGoogleChrome)
        self.ObjSeleniumConfiguration.CurrentWindowsHandler = Drivers.current_window_handle
        # Drivers.switch_to.alert.accept()
        self.FrameworkInitializer.Driver = Drivers
        self.ObjSeleniumConfiguration.IsAppium=False


    def switchingtomobileapp(self):
        self.FrameworkInitializer.AppiumStartUp_Remote()



    #Code for dashboard
    def Dashboard_depricated(self):
        try:
            time.sleep(4)
            #Get Locators for Gateway Connectivity Status
            print("*************gateway Connectivity Status************************")
            gtwyConnectivityStatus_Parameters = ["Gateway Connectivity Status - Total",
                                                "Gateway Connectivity Status - Connected",
                                                "Gateway Connectivity Status - Disconnected",
                                                "Gateway Connectivity Status - Temporary Disconnected",
                                                "Gateway Connectivity Status (More than 12 hours) - Disconnected Total",
                                                "Gateway Connectivity Status (More than 12 hours) - Disconnected due to Power Failure",
                                                "Gateway Connectivity Status (More than 12 hours) - Disconnected due to No Wan",
                                                "Gateway Connectivity Status (Less than 12 hours) - Temporary Disconnected Total",
                                                "Gateway Connectivity Status (Less than 12 hours) - Temporary Disconnected Power Failure",
                                                "Gateway Connectivity Status (Less than 12 hours) -  Temporary Disconnected No Wan"
                                                ]

            for x in range(len(gtwyConnectivityStatus_Parameters)):

                Locate = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage", "gatewayConnectivityStatus",
                                                              gtwyConnectivityStatus_Parameters[x])

                print("Locator Type = " +str(Locate[0]))
                print("Locator Value = " + str(Locate[1]))
                print("Locator Index = " + str(Locate[2]))

                try:
                    gtwyConnectivityStatus_Web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(Locate[1]).text
                except Exception as e:
                    print("Element not found")
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)

                #print(gtwyConnectivityStatus_Web)
                path = self.ObjSeleniumConfiguration.TestDataFolderPath + "\\" + "Insight_Dashboard.xlsx"
                path = self.ObjSeleniumConfiguration.TestDataFolderPath + "\\" + "Insight_Dashboard.xlsx"
                print(path)
                connectionString = self.FrameworkInitializer.GetTestDataFromDataTable("Connection String")
                sqlQuery = self.FrameworkInitializer.GetTestDataFromDataTable("SQL Query")

                #execute connection string
                import pyodbc
                conn = pyodbc.connect(connectionString)
                cursor = conn.cursor()
                print("sqlQuery is = " + sqlQuery)
                cursor.execute(sqlQuery)
                conn.commit()
                rowcount = cursor.rowcount
                #print("Expected Row count is = " + expectedRowCount)
                print("[Info]: Row Count is = " + str(rowcount))
                counter = 0
                dataColumnMatch = self.FrameworkInitializer.GetTestDataFromDataTable("dataColumnMatch")
                print("*******************456456" +str(dataColumnMatch))
                if dataColumnMatch == "NA":
                    pass
                else:
                    for row in cursor:
                        counter = counter + 1
                        print("Data column match value = " + str(dataColumnMatch))
                        dataColumnMatch = int(dataColumnMatch)
                        print((row[dataColumnMatch]))
                        #print("[Details: Info]: Gateway Serial No. is = " + str(row[dataColumnMatch]))
                        #actualDatafromQuery = str(row[dataColumnMatch])


                from openpyxl import load_workbook
                workbook = load_workbook(filename=path)
                sheet = workbook.active
                x  = x+2
                g1 = sheet.cell(row=self.ObjSeleniumConfiguration.intCurrentIteration, column=7)
                g1.value = gtwyConnectivityStatus_Web
                workbook.save(filename=path)


            #Get Locators for Critical Event
            print("**********************Critical Events************************")
            gtwyRTCFail = ["Gateway Critical Event - RTC Fail"]
            for x in range(len(gtwyRTCFail)):

                Locate = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage", "gatewayCriticalEvents",
                                                              gtwyRTCFail[x])


                print("Locator Type = " +str(Locate[0]))
                print("Locator Value = " + str(Locate[1]))
                print("Locator Index = " + str(Locate[2]))
                try:
                    gtwyRTCFail_Web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(Locate[1]).text
                except Exception as e:
                    print("Element not found")
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, fname, exc_tb.tb_lineno)
                path = self.ObjSeleniumConfiguration.TestDataFolderPath + "\\" + "Insight_Dashboard.xlsx"
                print(path)
                from openpyxl import load_workbook
                workbook = load_workbook(filename=path)
                sheet = workbook.active
                x = x + 12
                g1 = sheet.cell(row=x, column=7)
                g1.value = gtwyRTCFail_Web
                g1.value = gtwyRTCFail_Web
                workbook.save(filename=path)

            #Get Locators for Meter Reading Status
            print("**********************Meter reading status************************")
            meterReadingStatus_Parameters = ["Meter Reading Status  - Total",
                                 "Meter Reading Status  - Read",
                                 "Meter Reading Status - UnRead",
                                 "Meter Reading Status - Temporary UnRead",
                                 "Meter Reading Status - UnRead More than One Day",
                                 "Meter Reading Status - Never Read",
                                 "Meter Reading Status - Temporary UnRead"]

            for x in range(len(meterReadingStatus_Parameters)):
                Locate = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage", "meterReadingStatus",meterReadingStatus_Parameters[x])

                print("Locator Type = " +str(Locate[0]))
                print("Locator Value = " + str(Locate[1]))
                print("Locator Index = " + str(Locate[2]))

                path = self.ObjSeleniumConfiguration.TestDataFolderPath + "\\" + "Insight_Dashboard.xlsx"
                print(path)
                from openpyxl import load_workbook
                workbook = load_workbook(filename=path)
                sheet = workbook.active
                x = x + 13
                g1 = sheet.cell(row=x, column=7)
                g1.value = "Vivek2"
                workbook.save(filename=path)

            #get Locators for Last read Critical Events
            print("**********************Last read Critical Events************************")
            lastReadCriticalEvents = ["Critical Events - Voltage Miss on Any Phase",
                                        "Critical Events - Current Miss on Any Phase",
                                        "Critical Events - Magnet Interface",
                                        "Critical Events - Neutral Disturbance",
                                        "Critical Events - High Temperature"]
            for x in range(len(lastReadCriticalEvents)):
                Locate = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage", "lastReadCriticalEvents",
                                                              lastReadCriticalEvents[x])

                print("Locator Type = " +str(Locate[0]))
                print("Locator Value = " + str(Locate[1]))
                print("Locator Index = " + str(Locate[2]))

                path = self.ObjSeleniumConfiguration.TestDataFolderPath + "\\" + "Insight_Dashboard.xlsx"
                print(path)
                from openpyxl import load_workbook
                workbook = load_workbook(filename=path)
                sheet = workbook.active
                x = x + 20
                g1 = sheet.cell(row=x, column=7)
                g1.value = "Vivek3"
                workbook.save(filename=path)

        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

    def etlExecution(self):
        import clr

        try:
            clr.AddReference(r"C:\Users\44454\Desktop\Python Working Framework\Release\ETLOnDemandRunner")
            import ETLOnDemandRunner
        except Exception as e:
            print(str(e))
            # Making Class Object
        obj = ETLOnDemandRunner.RunETL()
        # get current time
        from datetime import datetime
        now = datetime.now()
        current_time_1 = now.strftime("%H:%M:%S")
        print("Current Time 1=", current_time_1)
        obj.Run("SSPMetercountByGateway", "Server=SMLI05194;Database=master;User Id=sa;Password=sa123;")
        # get current time
        from datetime import datetime
        now = datetime.now()
        current_time_2 = now.strftime("%H:%M:%S")
        print("Current Time 2=", current_time_2)
        FMT = '%H:%M:%S'
        tdelta = datetime.strptime(current_time_2, FMT) - datetime.strptime(current_time_1, FMT)
        print("Time taken to execute etl = " + str(tdelta))
        # Importing the library
        import os

        import psutil

        # Calling psutil.cpu_precent() for 4 seconds
        print('The CPU usage is: ', psutil.cpu_percent(4))
        # Getting % usage of virtual_memory ( 3rd field)
        print('RAM memory % used:', psutil.virtual_memory()[2])

    #Code to compare the dashboard with pre condition already set,
    def Dashboard(self):
        time.sleep(4)
        totalIteration = self.ObjSeleniumConfiguration.IterationTo
        # for x in range(totalIteration):
        totalKeywords = self.FrameworkInitializer.GetTestDataFromDataTable("Test Scenario")
        expectedData = self.FrameworkInitializer.GetTestDataFromDataTable("Expected Data")
        Locate = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus",
                                                      totalKeywords)
        try:
            gatewayAndMeterDetails = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(Locate[1]).text
            print("[Info]: " + str(totalKeywords) + " = " + str(gatewayAndMeterDetails))
        except Exception as e:
            print("Element not found")
            self.ObjSeleniumConfiguration.TestScenarioName = str(totalKeywords)
            test_check.writeResultExclReport(self, actualData="Element Not Found",
                                             dataToVerify="Verification Failed",
                                             result="Fail")
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        if gatewayAndMeterDetails == str(expectedData):
            print("Pass")
            test_check.writeResultExclReport(self, actualData=gatewayAndMeterDetails,
                                             dataToVerify=expectedData,
                                             result="Pass")
        else:
            print("Fail")
            test_check.writeResultExclReport(self, actualData=gatewayAndMeterDetails,
                                             dataToVerify=expectedData,
                                             result="Fail")






#Code for dashboard
    def Dashboard_One(self):
        try:
            self.FrameworkCommonLibrary.Driver.refresh()
            time.sleep(4)
            totalIteration = self.ObjSeleniumConfiguration.IterationTo
            #for x in range(totalIteration):
            totalKeywords = self.FrameworkInitializer.GetTestDataFromDataTable("Test Scenario")
            Locate = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus",totalKeywords)
            try:
                #element = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(Locate[1])
                #self.FrameworkCommonLibrary.Driver.execute_script("arguments[0].scrollIntoView();", element)
                #time.sleep(3)


                gatewayAndMeterDetails = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(Locate[1]).text
                print("[Info]: " +str(totalKeywords) +" = " + str(gatewayAndMeterDetails))
            except Exception as e:
                print("Element not found")
                self.ObjSeleniumConfiguration.TestScenarioName = str(totalKeywords)
                test_check.writeResultExclReport(self, actualData="Element Not Found",
                                                 dataToVerify="Verification Failed",
                                                 result="Fail")
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)

            path = self.ObjSeleniumConfiguration.TestDataFolderPath + "\\" + "Insight_Dashboard.xlsx"
            connectionString = self.FrameworkInitializer.GetTestDataFromDataTable("Connection String")
            sqlQuery = self.FrameworkInitializer.GetTestDataFromDataTable("SQL Query")
            # execute connection string
            import pyodbc
            conn = pyodbc.connect(connectionString)
            cursor = conn.cursor()
            cursor.execute(sqlQuery)
            conn.commit()
            rowcount = cursor.rowcount
            print("[Info]: Row Count is = " + str(rowcount))
            counter = 0
            databaseColumn = self.FrameworkInitializer.GetTestDataFromDataTable("Database Column Match")
            #print(type(databaseColumn))
            dataFromDatabase = ''
            try:
                databaseColumn = int(databaseColumn)
            except Exception as e:
                print("String or other conversion inititated")
            if str(databaseColumn).upper() == "NO":
                pass
            else:
                for row in cursor:
                    counter = counter + 1
                    dataFromDatabase = row[databaseColumn]
                    print("Data Fetched From Database for " +str(totalKeywords) + " = " +str((row[databaseColumn])))


            if str(dataFromDatabase) ==  str(gatewayAndMeterDetails):
                print("Pass")
                self.ObjSeleniumConfiguration.TestScenarioName = str(totalKeywords)
                test_check.writeResultExclReport(self, actualData=gatewayAndMeterDetails,
                                                 dataToVerify=dataFromDatabase,
                                                 result="Pass")
            else:
                print("Fail")
                self.ObjSeleniumConfiguration.TestScenarioName = str(totalKeywords)
                test_check.writeResultExclReport(self, actualData=gatewayAndMeterDetails,
                                                 dataToVerify=dataFromDatabase,
                                                 result="Fail")
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)


    def dashboardValuesSum(self):
        #get locators from OR for gateway
        #Pie Chart
        time.sleep(10)
        totalGateway_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus","Gateway Connectivity Status - Total")
        print(totalGateway_locator)
        temporaryDisconnectedGateway_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus","Gateway Connectivity Status - Temporary Disconnected")
        disconnectedGateway_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus","Gateway Connectivity Status - Disconnected")
        connectedGateway_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus","Gateway Connectivity Status - Connected")

        #DisconnectedMorethan12Hours
        totalDisconnectedGateway_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus","Gateway Connectivity Status (More than 12 hours) - Disconnected Total")
        disconnectedPowerFailure_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus","Gateway Connectivity Status (More than 12 hours) - Disconnected due to Power Failure")
        disconnectedNoWan_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus","Gateway Connectivity Status (More than 12 hours) - Disconnected due to No Wan")

        #temporary disconnected
        tempTotalDisconnected_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus","Gateway Connectivity Status (Less than 12 hours) - Temporary Disconnected Total")
        tempPowerFailure_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus","Gateway Connectivity Status (Less than 12 hours) - Temporary Disconnected Power Failure")
        tempNoWan_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus","Gateway Connectivity Status (Less than 12 hours) -  Temporary Disconnected No Wan")





        #Getting elements from the webpage
        try:
            totalGateway= self.FrameworkCommonLibrary.Driver.find_element_by_xpath(totalGateway_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        try:
            temporaryDisconnectedGateway= self.FrameworkCommonLibrary.Driver.find_element_by_xpath(temporaryDisconnectedGateway_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
        try:
            disconnectedGateway= self.FrameworkCommonLibrary.Driver.find_element_by_xpath(disconnectedGateway_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
        try:
            connectedGateway = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(connectedGateway_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
        try:
            totalDisconnectedGateway = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(totalDisconnectedGateway_locator[1]).text
            totalDisconnectedGateway = int(totalDisconnectedGateway)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
        try:
            disconnectedPowerFailure = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(disconnectedPowerFailure_locator[1]).text
            disconnectedPowerFailure = int(disconnectedPowerFailure)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
        try:
            disconnectedNoWan = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(disconnectedNoWan_locator[1]).text
            disconnectedNoWan = int(disconnectedNoWan)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
        try:
            tempTotalDisconnected = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(tempTotalDisconnected_locator[1]).text
            tempTotalDisconnected = int(tempTotalDisconnected)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
        try:
            tempPowerFailure = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(tempPowerFailure_locator[1]).text
            tempPowerFailure = int(tempPowerFailure)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
        try:
            tempNoWan = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(tempNoWan_locator[1]).text
            tempNoWan = int(tempNoWan)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)




        #Validating the logic
        #if totalGateway == (temporaryDisconnectedGateway+disconnectedGateway+connectedGateway):
            #test_check.writeResultExclReport(self, actualData=totalGateway,
                                             #dataToVerify=temporaryDisconnectedGateway+disconnectedGateway+connectedGateway,
                                             #result="Pass")
        #else:
            #test_check.writeResultExclReport(self, actualData=totalGateway,
                                             #dataToVerify=temporaryDisconnectedGateway + disconnectedGateway,
                                             #result="Fail")
        if totalDisconnectedGateway == (disconnectedPowerFailure+disconnectedNoWan):
            test_check.writeResultExclReport(self, actualData=totalDisconnectedGateway,
                                             dataToVerify=disconnectedPowerFailure+disconnectedNoWan,
                                             result="Pass")
        else:
            test_check.writeResultExclReport(self, actualData=totalDisconnectedGateway,
                                             dataToVerify=disconnectedPowerFailure + disconnectedNoWan,
                                             result="Fail")
        if tempTotalDisconnected == (tempPowerFailure+tempNoWan):
            test_check.writeResultExclReport(self, actualData=tempTotalDisconnected,
                                             dataToVerify=tempPowerFailure+tempNoWan,
                                             result="Pass")
        else:
            test_check.writeResultExclReport(self, actualData=tempTotalDisconnected,
                                             dataToVerify=tempPowerFailure + tempNoWan,
                                             result="Fail")


        #getting the sum of meter details card




    def storedProcedureExecution(self):
        import psycopg2
        try:
            testScenario = self.FrameworkInitializer.GetTestDataFromDataTable("Test Scenario")
            self.ObjSeleniumConfiguration.TestScenarioName = str(testScenario)
            serverIP = self.FrameworkInitializer.GetTestDataFromDataTable("Host(Server IP)")
            port = self.FrameworkInitializer.GetTestDataFromDataTable("Port")
            database = self.FrameworkInitializer.GetTestDataFromDataTable("Database")
            user = self.FrameworkInitializer.GetTestDataFromDataTable("User")
            password = self.FrameworkInitializer.GetTestDataFromDataTable("Password")
            storedProcedure = self.FrameworkInitializer.GetTestDataFromDataTable("Stored Procedure")
            print(storedProcedure)
            connect = psycopg2.connect(host =serverIP,port=port,database=database,user=user,password=password)
            cursor = connect.cursor()
            cursor.execute(storedProcedure)
            connect.commit()
            test_check.writeResultExclReport(self, actualData="Procedure Execution Successful",
                                             dataToVerify="Procedure Execution",
                                             result="Pass")
            self.FrameworkCommonLibrary.Driver.refresh()
            time.sleep(4)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            test_check.writeResultExclReport(self, actualData="Error in Procedure Execution",
                                             dataToVerify="Procedure Execution",
                                             result="Fail")

    def PageRefresh(self):
        self.FrameworkCommonLibrary.Driver.refresh()
        time.sleep(6)


    #code to validated the detailed page
    def validateDashboardDetails(self):
        try:

            #get the testdata from excel
            testScenario = self.FrameworkInitializer.GetTestDataFromDataTable("Page Details")
            cardDetails = self.FrameworkInitializer.GetTestDataFromDataTable("Card Detail")
            pageTitle_testData = self.FrameworkInitializer.GetTestDataFromDataTable("Verify Page Title")
            firstDropdown = self.FrameworkInitializer.GetTestDataFromDataTable("Validate First Dropdown Details")
            secondDropdown = self.FrameworkInitializer.GetTestDataFromDataTable("Validate Second Dropdown Details")
            self.ObjSeleniumConfiguration.TestScenarioName = testScenario

            #Get the values of locators from OR

            failureTypeValue_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus",testScenario)
            pageTitle_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page","dashboardDetailsPage","pageTitle")
            logoutLinkBtnId = "userlink_btn"
            logoutlinkXpath = "//a[contains(text(),'Logout')]"
            
            # lOCATORS for gateway details card
            if cardDetails.upper() == "GATEWAYDETAILS":
                reason_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page", "dashboardDetailsPage","reason")
                duration_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page", "dashboardDetailsPage","duration")
                pageTitle_2_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page","dashboardDetailsPage", "pageTitle_two")
                #viewButton_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page","dashboardDetailsPage","viewButton")
                #print(viewButton_locator)
                
            # locators for gateway critical event card
            elif cardDetails.upper() == "GATEWAYCRITICALEVENTS":
                reason_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page", "dashboardDetailsPage","reason")
                duration_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page", "dashboardDetailsPage","duration")
                pageTitle_2_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page","dashboardDetailsPage", "pageTitle_two_gtwyCriticalEvents")

            # locators for meter reading
            elif cardDetails.upper() == "METERDETAILS":
                reason_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page", "dashboardDetailsPage", "reason_meterDetails")
                duration_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page", "dashboardDetailsPage","duration")
                pageTitle_2_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page","dashboardDetailsPage", "pageTitle_two_meterDetails")

            elif cardDetails.upper() == "METERCRITICALEVENTS":
                reason_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page", "dashboardDetailsPage","reason_meterCriticalEvents")
                duration_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page", "dashboardDetailsPage","duration")
                pageTitle_2_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page","dashboardDetailsPage", "pageTitle_two_meterCriticalEvents")

            # Locaating the elements on the webpage
            try:
                WebDriverWait(self.FrameworkCommonLibrary.Driver, 100).until(EC.presence_of_element_located((By.XPATH, failureTypeValue_locator[1])))
                failureTypeValue = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(failureTypeValue_locator[1]).text
                #print("***************failuretypevalue" +str(failureTypeValue))

            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)
                print("Element not found")

            # Click on the element and navigate to the respective detailed page
            self.FrameworkCommonLibrary.Driver.find_element_by_xpath(failureTypeValue_locator[1]).click()
            WebDriverWait(self.FrameworkCommonLibrary.Driver, 100).until(EC.presence_of_element_located((By.ID, pageTitle_locator[1])))

            #Locate the elements
            try:
                pageTitleValue = self.FrameworkCommonLibrary.Driver.find_element_by_id(pageTitle_locator[1]).text
                self.FrameworkCommonLibrary.Driver.refresh()
                time.sleep(5)
                pageTitleTwoValue = self.FrameworkCommonLibrary.Driver.find_element_by_id(pageTitle_2_locator[1]).text
                print(pageTitle_2_locator[1])
                print(pageTitleTwoValue)

            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)
                print("Element not found")

            #Validating the results
            #page title
            if pageTitleValue.upper() == pageTitle_testData.upper():
                test_check.writeResultExclReport(self, actualData=pageTitleValue,
                                                 dataToVerify=pageTitle_testData,
                                                 result="Pass")
                print("Page title validation passed")
            else:
                test_check.writeResultExclReport(self, actualData=pageTitleValue,
                                                 dataToVerify=pageTitle_testData,
                                                 result="Fail")
                print("Page title validation failed")

            # first dropdown
            if (firstDropdown).upper() == "NOT APPLICABLE":
                pass
            else:
                print(self.FrameworkInitializer.GetTestDataFromDataTable("Validate First Dropdown Details"))
                failureDropdown = Select(self.FrameworkCommonLibrary.Driver.find_element_by_id(reason_locator[1]))
                failureDropdown.select_by_visible_text(firstDropdown)
                if (failureDropdown.first_selected_option).text == firstDropdown:
                    test_check.writeResultExclReport(self, actualData=(failureDropdown.first_selected_option).text,
                                                     dataToVerify=firstDropdown,
                                                     result="Pass")
                    print("Failure type validation passed")
                else:
                    test_check.writeResultExclReport(self, actualData=(failureDropdown.first_selected_option).text,
                                                     dataToVerify=firstDropdown,
                                                     result="Fail")
                    print("Failure validation failed")

            # Validation time duration # Second dropdown
            if (secondDropdown).upper() == "NOT APPLICABLE":
                pass
            else:
                print(self.FrameworkInitializer.GetTestDataFromDataTable("Validate Second Dropdown Details"))
                reasonDropdown = Select(self.FrameworkCommonLibrary.Driver.find_element_by_id(duration_locator[1]))
                reasonDropdown.select_by_visible_text(secondDropdown)
                viewButton_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page","dashboardDetailsPage","viewButton")
                #print(viewButton_locator)
                #print(viewButton_locator[1])
                time.sleep(2)
                self.FrameworkCommonLibrary.Driver.find_element_by_id(viewButton_locator[1]).click()
                time.sleep(7)
                if (reasonDropdown.first_selected_option).text == secondDropdown:
                    test_check.writeResultExclReport(self, actualData=(reasonDropdown.first_selected_option).text,
                                                     dataToVerify=secondDropdown,
                                                     result="Pass")
                    print("Time duration validation passed")
                else:
                    test_check.writeResultExclReport(self, actualData=(reasonDropdown.first_selected_option).text,
                                                     dataToVerify=secondDropdown,
                                                     result="Fail")
                    print("Time duration validation failed")

            #Validating for the title 2
            #split the text
            WebDriverWait(self.FrameworkCommonLibrary.Driver, 100).until(EC.presence_of_element_located((By.ID, pageTitle_2_locator[1])))

            numberOfGatewayDisplayed = pageTitleTwoValue.split()
            #print("**********************numberofGateay displayued page title"+str(numberOfGatewayDisplayed))
            integerValueofGtwyDisplaced = str(numberOfGatewayDisplayed[0])
            print(integerValueofGtwyDisplaced)
            print(failureTypeValue)
            if integerValueofGtwyDisplaced == failureTypeValue:
                test_check.writeResultExclReport(self, actualData=failureTypeValue,
                                                 dataToVerify=integerValueofGtwyDisplaced,
                                                 result="Pass")
                print("Gateway Count validation passed")
            else:
                test_check.writeResultExclReport(self, actualData=failureTypeValue,
                                                 dataToVerify=integerValueofGtwyDisplaced,
                                                 result="Fail")
                print("Gateway Count validation Failed")
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        #Code to export data to csv
        test_check.exportDataFromDatabasetoExcel(self)

        time.sleep(5)
        # Code to download the csv and compare the same
        downloadCSV_ORValue = self.FrameworkCommonLibrary.Locators("Dashboard_Detailed_Page","dashboardDetailsPage", "downloadCSV")
        self.FrameworkCommonLibrary.Driver.find_element_by_xpath(downloadCSV_ORValue[1]).click()

        self.FrameworkCommonLibrary.Driver.find_element_by_id(logoutLinkBtnId).click()
        time.sleep(2)
        self.FrameworkCommonLibrary.Driver.find_element_by_xpath(logoutlinkXpath).click()
        
        # wait till download has finished
        time.sleep(10)

        #Code to compare csv
        test_check.compareCSV(self)



    #Code to export data from sql to excel and csv for matching and also to database sheet
    def exportDataFromDatabasetoExcel(self):
        # Replace String with the condition
        import json
        try:

            substring = self.FrameworkInitializer.GetTestDataFromDataTable("Replaceable String")
            substring = json.loads(substring)
            key_list = list(substring.keys())
            val_list = list(substring.values())

            Query = self.FrameworkInitializer.GetTestDataFromDataTable("SQL Query")
            fullstring = str(Query)

            for x in range(len(key_list)):
                if key_list[x] in fullstring:
                    print("Found!")
                    fullstring = fullstring.replace(key_list[x], val_list[x])
                else:
                    print("Not Found")
            print(fullstring)

            connectionString = str(self.FrameworkInitializer.GetTestDataFromDataTable("Connection String"))
            Query = fullstring

            splitedConnectionString = connectionString.split('||')
            print(len(splitedConnectionString))
            splittedQuery = Query.split('||')
            print(len(splittedQuery))
            frames = []
            for x in range(len(splitedConnectionString)):
                conn = pyodbc.connect(splitedConnectionString[x])
                sql_query = pd.read_sql_query(splittedQuery[x], conn)
                df = pd.DataFrame(sql_query)
                frames.append(df)
            result = pd.concat([pd.concat((frames), axis=1, join="inner")])
            # result = pd.concat([pd.concat((frames))], axis=1, join="inner")
            result.to_excel(r'..\..\..\exported_data_excel.xlsx', index=False)
            result.to_csv(r"..\..\..\exported_data_excel.csv", index=False)

            # copy the file to the database sheet
            excelDatabaseSheetName = self.FrameworkInitializer.GetTestDataFromDataTable("Export Data to Excel")
            path = r'..\\SeleniumAutomation\\DataBase\\' +excelDatabaseSheetName
            df.to_excel(path, index=False)

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)




    # Function Depricated
    def exportDataFromDatabasetoExcel1(self):
        # Replace String with the condition
        try:
            substring = "XXX"
            Query = self.FrameworkInitializer.GetTestDataFromDataTable("SQL Query")
            fullstring = str(Query)
            replaceStringDetails = str(self.FrameworkInitializer.GetTestDataFromDataTable("Condition"))
            if substring in fullstring:
                print("Found!")
                Query = fullstring.replace("XXX", replaceStringDetails)
                print(fullstring.replace("XXX", replaceStringDetails))

            connectionString = str(self.FrameworkInitializer.GetTestDataFromDataTable("Connection String"))
            #conn = pyodbc.connect(connectionString)
            conn = pyodbc.connect(connectionString)
            #sql_query = pd.read_sql_query(self.FrameworkInitializer.GetTestDataFromDataTable("Connection String"),conn)
            sql_query = pd.read_sql_query(Query,conn)
            print(Query)
            df = pd.DataFrame(sql_query)
            df.to_excel(r'..\..\..\exported_data_excel.xlsx', index=False)

            # Convert the downloaded excel to csv
            read_file = pd.read_excel(r'..\..\..\exported_data_excel.xlsx')
            read_file.to_csv(r"..\..\..\exported_data_excel.csv",index=None,header=True)

            # copy the file to the database sheet
            excelDatabaseSheetName = self.FrameworkInitializer.GetTestDataFromDataTable("Export Data to Excel")
            path = r'..\\SeleniumAutomation\\DataBase\\' +excelDatabaseSheetName
            df.to_excel(path, index=False)


        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

    # Code to compare two csv
    def compareCSV(self):
        try:
            # code to get the last downloaded file and lst modified time
            folder_path = r'C:\Users\45125\Downloads'
            file_type = '\*csv'
            files = glob.glob(folder_path + file_type)
            max_file = max(files, key=os.path.getctime)
            print(max_file)

            # get the last modified time of the file
            fileStatsObj = os.stat(max_file)
            modificationTime = time.ctime(fileStatsObj[stat.ST_MTIME])
            print("Last Modified Time : ", modificationTime)

            # compare csv
            df1 = pd.read_csv(r"..\..\..\exported_data_excel.csv")
            df2 = pd.read_csv(max_file)
            df1.sort_values(df1.columns[0], axis=0, inplace=True)
            df2.sort_values(df2.columns[0], axis=0, inplace=True)
            print(df1)
            print(df2)
            
            # comparing headers
            if len(df1.columns) == len(df2.columns):
                if ((df1.axes[1]==df2.axes[1]).any()):  # ==> 1 is to identify columns
              
                    print("Headers Matched")
                    test_check.writeResultExclReport(self,
                                                     actualData="Headers Matched",
                                                     dataToVerify="Headers Matched",
                                                     result="Pass")
               
                    if df1.empty and df2.empty:
                        print("Please check Empty values are identified in the exported Dataset")
                        test_check.writeResultExclReport(self,
                                                         actualData="Please check Empty values are identified in the exported Dataset - Name of file used for matching = " + str(
                                                             max_file),
                                                         dataToVerify="Last modified time of the file was = " + str(
                                                             modificationTime),
                                                         result="Pass")
                    else:
                        comparison_values = df1.values == df2.values
                        print(type(comparison_values))
                        print(comparison_values)
                        #exists = (False == comparison_values)
                        exists = False in comparison_values
                        print(exists)
                        if str(exists).upper() == "TRUE":
                            ######################################Code to find the index where data matching failed###################
                            ''' Get index positions of value in dataframe i.e. dfObj.'''
                            listOfPos = list()
                            # Get bool dataframe with True at positions where the given value exists
                            df = pd.DataFrame(comparison_values)
                            print(df)
                            result = df.isin([False])
                            print(result)
                            # Get list of columns that contains the value
                            seriesObj = result.any()
                            columnNames = list(seriesObj[seriesObj == True].index)
                            # Iterate over list of columns and fetch the rows indexes where value exists
                            for col in columnNames:
                                rows = list(result[col][result[col] == True].index)
                                for row in rows:
                                    listOfPos.append((row, col))

                            ##########################################################################################################
                            print("Data Matching Failed")
                            test_check.writeResultExclReport(self,
                                                             actualData="Name of file used for matching = " + str(max_file) +"Index Location = " +str(listOfPos),
                                                             dataToVerify="Last modified time of the file was = " + str(
                                                                 modificationTime),
                                                             result="Fail")

                        else:
                            print("Data Matching Passed")
                            test_check.writeResultExclReport(self,
                                                             actualData="Name of file used for matching = " + str(max_file),
                                                             dataToVerify="Last modified time of the file was = " + str(
                                                                 modificationTime),
                                                             result="Pass")

                else:
                    print("Headers Mismatched")
                    test_check.writeResultExclReport(self,
                                                     actualData="Headers Mismatched",
                                                     dataToVerify="Headers Mismatched",
                                                     result="Fail")
            else:
                print("Column Size Mismatched")
                test_check.writeResultExclReport(self,
                                                 actualData="Headers of the compared CSV is Not Matching!!",
                                                 dataToVerify="Headers of the compared CSV should Match",
                                                 result="Fail")





        except Exception as e:
            print("Test Failed")
            print(str(e))
            test_check.writeResultExclReport(self, actualData="Error in matching file " + str(max_file),
                                             dataToVerify="Last modified time of the file was = " + str(
                                                 modificationTime),
                                             result="Fail")
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

    def readPieChart_GatewayDetails(self):
        time.sleep(5)
        print("Pie Chart validation execution Started")
        data = self.FrameworkCommonLibrary.Driver.execute_script("return  $('#jqxChart').jqxChart('getInstance').source")
        connectionString = self.FrameworkInitializer.GetTestDataFromDataTable("Connection String")
        sqlQuery = self.FrameworkInitializer.GetTestDataFromDataTable("SQL Query")
        databaseColumnMatch = self.FrameworkInitializer.GetTestDataFromDataTable("Database Column Match")
        testScenario = self.FrameworkInitializer.GetTestDataFromDataTable("Test Scenario")

        for x in range(len(data)):
            Status = data[x].get('Status')
            Value = data[x].get('Value')
            #print(str(Status) + '=' + str(Value))

            if testScenario.upper() == "GATEWAY CONNECTIVITY STATUS - CONNECTED":
                self.ObjSeleniumConfiguration.TestScenarioName = "GATEWAY CONNECTIVITY STATUS - CONNECTED"
                valueFromDatabase = test_check.getvaluesFromDatabase(self, ConnectionString=connectionString,Query=sqlQuery,DataMatch=databaseColumnMatch )
                #print("str(valueFromDatabase) == str(Value)")
                if Status.upper() == "CONNECTED":
                    if str(valueFromDatabase) == str(Value):
                        print("[Result]: Pass : GATEWAY CONNECTIVITY STATUS - CONNECTED")
                        test_check.writeResultExclReport(self, actualData=str(Value), dataToVerify=valueFromDatabase,
                                                 result="Pass")
                    else:
                        print("[Result]: Fail : GATEWAY CONNECTIVITY STATUS - CONNECTED")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Fail")



            if testScenario.upper() == "GATEWAY CONNECTIVITY STATUS - DISCONNECTED":
                self.ObjSeleniumConfiguration.TestScenarioName = "GATEWAY CONNECTIVITY STATUS - DISCONNECTED"
                valueFromDatabase = test_check.getvaluesFromDatabase(self, ConnectionString=connectionString,Query=sqlQuery,DataMatch=databaseColumnMatch )
                #print("str(valueFromDatabase) == str(Value)")
                if Status.upper() == "DISCONNECTED":
                    if str(valueFromDatabase) == str(Value):
                        print("[Result]: Pass : GATEWAY CONNECTIVITY STATUS - DISCONNECTED")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Pass")
                    else:
                        print("[Result]: Fail : GATEWAY CONNECTIVITY STATUS - DISCONNECTED")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Fail")


            if testScenario.upper() == "GATEWAY CONNECTIVITY STATUS - TEMPORARY DISCONNECTED":
                self.ObjSeleniumConfiguration.TestScenarioName = "GATEWAY CONNECTIVITY STATUS - TEMPORARY DISCONNECTED"
                valueFromDatabase = test_check.getvaluesFromDatabase(self, ConnectionString=connectionString,
                                                                     Query=sqlQuery, DataMatch=databaseColumnMatch)


                print("GATEWAY CONNECTIVITY STATUS - TEMPORARY DISCONNECTED")

                if Status.upper() == "TEMPORARILY DISCONNECTED":
                    if str(valueFromDatabase) == str(Value):
                        print("[Result]: Pass : GATEWAY CONNECTIVITY STATUS - TEMPORARY DISCONNECTED")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Pass")

                    else:
                        print("[Result]: Fail : GATEWAY CONNECTIVITY STATUS - TEMPORARY DISCONNECTED")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Fail")



    def readPieChart_MeterDetails(self):
        time.sleep(5)
        print("Pie Chart validation execution Started")
        data = self.FrameworkCommonLibrary.Driver.execute_script("return  $('#jqxmeterChart').jqxChart('getInstance').source")
        connectionString = self.FrameworkInitializer.GetTestDataFromDataTable("Connection String")
        sqlQuery = self.FrameworkInitializer.GetTestDataFromDataTable("SQL Query")
        databaseColumnMatch = self.FrameworkInitializer.GetTestDataFromDataTable("Database Column Match")
        testScenario = self.FrameworkInitializer.GetTestDataFromDataTable("Test Scenario")

        #get locators and their text values from the dashboard page.
        totalMeters_locators = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One","gatewayConnectivityStatus","Meter Reading Status  - Total")
        unreadMoreThanOneDay_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus", "Meter Reading Status - UnRead More than One Day")
        neverReadMoreThanOneDay_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus", "Meter Reading Status - Never Read More than One Day")
        temporarilyUnread_locator = self.FrameworkCommonLibrary.Locators("Dashboard_Homepage_One", "gatewayConnectivityStatus", "Meter Reading Status - Temporary UnRead More than One Day")

        #getting text values of the locators
        totalMeters_value = int(self.FrameworkCommonLibrary.Driver.find_element_by_xpath(totalMeters_locators[1]).text)
        unreadMoreThanOneDay_value = int(self.FrameworkCommonLibrary.Driver.find_element_by_xpath(unreadMoreThanOneDay_locator[1]).text)
        neverReadMoreThanOneDay_value = int(self.FrameworkCommonLibrary.Driver.find_element_by_xpath(neverReadMoreThanOneDay_locator[1]).text)
        temporarilyUnreadMoreThanOneDay_value = int(self.FrameworkCommonLibrary.Driver.find_element_by_xpath(temporarilyUnread_locator[1]).text)

        for x in range(len(data)):
            Status = data[x].get('Status')
            Value = data[x].get('value')
            print(str(Status) + '=' + str(Value))


            if testScenario.upper() == "METER READING STATUS  - READ":
                self.ObjSeleniumConfiguration.TestScenarioName = "METER READING STATUS  - READ"
                print("METER READING STATUS  - READ")
                valueFromDatabase = test_check.getvaluesFromDatabase(self, ConnectionString=connectionString,
                                                                     Query=sqlQuery, DataMatch=databaseColumnMatch)

                if Status.upper() == "READ":
                    if str(valueFromDatabase) == str(Value):
                        print("[Result]: Pass: READ")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Pass")
                    else:
                        print("[Result]: Fail: READ")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Fail")

            if testScenario.upper() == "METER READING STATUS - UNREAD":
                self.ObjSeleniumConfiguration.TestScenarioName = "METER READING STATUS - UNREAD"
                print("METER READING STATUS - UNREAD")
                valueFromDatabase = test_check.getvaluesFromDatabase(self, ConnectionString=connectionString,
                                                                     Query=sqlQuery, DataMatch=databaseColumnMatch)

                if Status.upper() == "UNREAD":
                    if str(valueFromDatabase) == str(Value):
                        print("[Result]: Pass: UNREAD")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Pass")
                        test_check.DashboardSumValues(self, totalMeters_value,int(Value),unreadMoreThanOneDay_value,neverReadMoreThanOneDay_value,temporarilyUnreadMoreThanOneDay_value)
                        # DashboardSumValues(self, TotalSum, *argv):
                    else:
                        print("[Result]: Fail:UNREAD ")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Fail")

            if testScenario.upper() == "METER READING STATUS - TEMPORARY UNREAD":
                self.ObjSeleniumConfiguration.TestScenarioName = "METER READING STATUS - TEMPORARY UNREAD"
                print("METER READING STATUS - TEMPORARY UNREAD")
                valueFromDatabase = test_check.getvaluesFromDatabase(self, ConnectionString=connectionString,
                                                                     Query=sqlQuery, DataMatch=databaseColumnMatch)

                if Status.upper() == "TEMPORARILY UNREAD":
                    if str(valueFromDatabase) == str(Value):
                        print("[Result]: Pass: ")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Pass")
                    else:
                        print("[Result]: Fail: TEMPORARILY UNREAD")
                        test_check.writeResultExclReport(self, actualData=str(Value),
                                                         dataToVerify=valueFromDatabase,
                                                         result="Fail")


    def getvaluesFromDatabase(self,ConnectionString,Query,DataMatch):
        import pyodbc
        conn = pyodbc.connect(ConnectionString)
        cursor = conn.cursor()
        #print("sqlQuery is = " + Query)
        cursor.execute(Query)
        conn.commit()
        counter = 0
        rowDataFetched = ''

        for row in cursor:
            counter = counter + 1
            #print("[Info]: Value = " + str(row[int(DataMatch)]))
            rowDataFetched = str(row[int(DataMatch)])
        return rowDataFetched


    def DashboardSumValues(self,TotalSum, *argv):
        sumvalue = int()
        for arg in argv:
            sumvalue = sumvalue + arg
            print(arg)
        if int(TotalSum) == int(sumvalue):
            test_check.writeResultExclReport(self, actualData=int(TotalSum),
                                                         dataToVerify=int(sumvalue),
                                                         result="Fail")
            #return True,sumvalue
        else:
            #return False,sumvalue
            test_check.writeResultExclReport(self, actualData=int(TotalSum),
                                             dataToVerify=int(sumvalue),
                                             result="Fail")

    def OnDemandMeterReading(self):
        import time
        try:
            print("On Demand Meter reading started")


            onDemandMeterLink = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand", "pageLink")
            self.FrameworkCommonLibrary.Driver.find_element_by_xpath(onDemandMeterLink[1]).click()

            pageTitle_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand", "pageTitle")
            WebDriverWait(self.FrameworkCommonLibrary.Driver, 100).until(EC.presence_of_element_located((By.ID, pageTitle_locator[1])))


            pageTitle_actual = self.FrameworkCommonLibrary.Driver.find_element_by_id(pageTitle_locator[1]).text
            pageTitle_expected = self.FrameworkInitializer.GetTestDataFromDataTable("Page Title")


            if str(pageTitle_actual) == str(pageTitle_expected):
                print("Page Title matched")
                test_check.writeResultExclReport(self, actualData=str(pageTitle_actual),
                                                 dataToVerify=str(pageTitle_expected),
                                                 result="Pass")
            else:
                print("Page Title Mismatched")
                test_check.writeResultExclReport(self, actualData=str(pageTitle_actual),
                                                 dataToVerify=str(pageTitle_expected),
                                                 result="Fail")



            Gateway_serialNumber =  self.FrameworkInitializer.GetTestDataFromDataTable("Meter_Serial_Number")


            serialnumberInputBox_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand", "Meter_Serial_Number")
            time.sleep(4)
            self.FrameworkCommonLibrary.Driver.find_element_by_id(serialnumberInputBox_locator[1]).send_keys(Gateway_serialNumber)


            submitBtn_locator=self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand", "Submit_button")
            self.FrameworkCommonLibrary.Driver.find_element_by_id(submitBtn_locator[1]).click()


            time.sleep(3)
            message_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand", "Expected Message")
            messageReceived = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(message_locator[1]).text
            expectedMessage = self.FrameworkInitializer.GetTestDataFromDataTable("Expected Message")

            if str(messageReceived) == str(expectedMessage):
                print("Message checked and passed")
                test_check.writeResultExclReport(self, actualData=str(messageReceived),
                                                 dataToVerify=str(expectedMessage),
                                                 result="Pass")
            else:
                print("Message Checking Failed")
                test_check.writeResultExclReport(self, actualData=str(messageReceived),
                                                 dataToVerify=str(expectedMessage),
                                                 result="Fail")
            time.sleep(2)


            ObBtn_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand", "OkBtn")
            btn = self.FrameworkCommonLibrary.Driver.find_element_by_id(ObBtn_locator[1])
            #print("*******************************" +str(ObBtn_locator[1]))
            self.FrameworkCommonLibrary.Driver.execute_script("arguments[0].click();", btn)

            time.sleep(4)
            self.FrameworkCommonLibrary.Driver.refresh()


            time.sleep(120)

            self.FrameworkCommonLibrary.Driver.refresh()
            time.sleep(10)

            # get the original details from database
            # Replace String with the condition
            import json
            substring = self.FrameworkInitializer.GetTestDataFromDataTable("Replaceable String")
            substring = json.loads(substring)
            key_list = list(substring.keys())
            val_list = list(substring.values())

            Query = self.FrameworkInitializer.GetTestDataFromDataTable("Sql Query")
            fullstring = str(Query)

            for x in range(len(key_list)):
                if key_list[x] in fullstring:
                    print("Found!")
                    fullstring = fullstring.replace(key_list[x], val_list[x])
                else:
                    print("Not Found")
            #print(fullstring)


            connectionString = self.FrameworkInitializer.GetTestDataFromDataTable("Connection String")
            sqlQuery = fullstring
            try:
                import pyodbc
                conn = pyodbc.connect(connectionString)
                cursor = conn.cursor()
                #print("sqlQuery is = " + sqlQuery)
                cursor.execute(sqlQuery)
                conn.commit()
            except Exception as e:
                print(str(e))
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)
            # ***************
            # print("Current Iteration is = " +str(self.ObjSeleniumConfiguration.intCurrentIteration))
            self.ObjSeleniumConfiguration.intCurrentIteration = self.ObjSeleniumConfiguration.intCurrentIteration + 1
            counter = 0
            rowDataFetched = []
            serialNumber_database = ''
            status_database = ''
            cumulativeEnergy_database = ''
            readingDate_database = ''
            lastBillingEnergy_database = ''
            lastBillingDemand_database = ''
            lastBillingDate_database = ''
            powerfactor_database = ''
            try:
                for row in cursor:
                    counter = counter + 1
                    print("[Info]: Database row imported = " + str(row))
                    print("**************************************************************")
                    rowDataFetched.append(str(row))
                    serialNumber_database = row[2]
                    status_database = row[1]
                    cumulativeEnergy_database = row[5]
                    # code to change the format of date
                    """dat = row[3]
                    dateAndTime = dat.split()
                    date = dateAndTime[0].split("-")
                    time = dateAndTime[1].split(":")
                    newDate = date[2] + "/" + date[1] + "/" + date[0]
                    newTime = time[0] + ":" + time[1]
                    #print(newDate)
                    #print(newTime)
                    newDateAndTime = newDate + " " + newTime
                    print(newDateAndTime)"""
                    readingDate_database = row[3]
                    lastBillingEnergy_database = row[6]
                    lastBillingDemand_database = row[7]
                    lastBillingDate_database = row[4]
                    powerfactor_database = row[8]
            except Exception as e:
                print(str(e))
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)



            databaseValues = [serialNumber_database,status_database,cumulativeEnergy_database,readingDate_database,lastBillingEnergy_database
                              ,lastBillingDemand_database,lastBillingDate_database,powerfactor_database]

            databaseValuesText = ["serialNumber", "status", "cumulativeEnergy", "readingDate","lastBillingEnergy", "lastBillingDemand",
                                  "lastBillingDate", "powerfactor"]
            serialNumber_web = ''
            status_database_web = ''
            cumulativeEnergy_database_web = ''
            readingDate_database_web = ''
            lastBillingEnergy_database_web = ''
            lastBillingDemand_database_web = ''
            lastBillingDate_database_web = ''
            powerfactor_database_web = ''


            try:
                serialNumber_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand",
                                                                            "Serial_Number")
                serialNumber_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                    serialNumber_locator[1]).text
            except Exception as e:
                print("Element not found")
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)




            try:
                status_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand", "Status")
                status_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(status_locator[1]).text
            except Exception as e:
                print("Element not found")
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)



            try:
                cumulativeEnergy_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand",
                                                                                "Cummulative_Energy")
                cumulativeEnergy_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                    cumulativeEnergy_locator[1]).text
            except Exception as e:
                print("Element not found")
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)



            try:
                readingDate_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand",
                                                                           "Reading_Date_Time")
                readingDate_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                    readingDate_locator[1]).text
            except Exception as e:
                print("Element not found")
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)



            try:
                lastBillingEnergy_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand",
                                                                                 "Last_billing_Energy")
                lastBillingEnergy_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                    lastBillingEnergy_locator[1]).text
            except Exception as e:
                print("Element not found")
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)


            try:
                lastBillingDemand_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand",
                                                                                 "Last_Billing_Demand")
                lastBillingDemand_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                    lastBillingDemand_locator[1]).text
            except Exception as e:
                print("Element not found")
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)



            try:
                lastBillingDate_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand",
                                                                               "Last_Billing_Date_Time")
                lastBillingDate_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                    lastBillingDate_locator[1]).text
            except Exception as e:
                print("Element not found")
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)


            try:
                powerfactor_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand",
                                                                           "Last_Billing_Power_Factor")
                powerfactor_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                    powerfactor_locator[1]).text
            except Exception as e:
                print("Element not found")
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)


            webValues = [serialNumber_web,status_database_web,cumulativeEnergy_database_web,readingDate_database_web,lastBillingEnergy_database_web,
                         lastBillingDemand_database_web,lastBillingDate_database_web,powerfactor_database_web]


            print(databaseValues)
            print(webValues)
            for x in range(len(databaseValues)):
                if str(databaseValues[x]) == str(webValues[x]):
                    
                    print(str(databaseValuesText[x])  + " : value matched = " + str(webValues[x]))
                    test_check.writeResultExclReport(self, actualData=str(databaseValuesText[x]) + " : " +str(webValues[x]),
                                                     dataToVerify=str(databaseValuesText[x]) + " : "  + str(databaseValues[x]),
                                                     result="Pass")
                else:
                    print(str(databaseValuesText[x])  + " : value Mismatched = " + str(webValues[x]))
                    test_check.writeResultExclReport(self, actualData=str(databaseValuesText[x]) + " : " + str(webValues[x]),
                                                     dataToVerify=str(databaseValuesText[x]) + " : " + str(
                                                         databaseValues[x]),
                                                     result="Fail")
        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

    def ViewMeterReading(self):
        print("View Meter reading started")


        onDemandMeterLink = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand", "pageLink")
        self.FrameworkCommonLibrary.Driver.find_element_by_xpath(onDemandMeterLink[1]).click()

        pageTitle_locator = self.FrameworkCommonLibrary.Locators("Meter_Reading_Page", "On_Demand", "pageTitle")
        WebDriverWait(self.FrameworkCommonLibrary.Driver, 100).until(EC.presence_of_element_located((By.ID, pageTitle_locator[1])))

        viewMeterReadingRadioBtn = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading", "View_Meter_Reading_button")
        self.FrameworkCommonLibrary.Driver.find_element_by_id(viewMeterReadingRadioBtn[1]).click()
        time.sleep(4)


        pageTitle_actual = self.FrameworkCommonLibrary.Driver.find_element_by_id(pageTitle_locator[1]).text
        pageTitle_expected = self.FrameworkInitializer.GetTestDataFromDataTable("Page Title")


        if str(pageTitle_actual) == str(pageTitle_expected):
            print("Page Title matched")
            test_check.writeResultExclReport(self, actualData=str(pageTitle_actual),
                                             dataToVerify=str(pageTitle_expected),
                                             result="Pass")
        else:
            print("Page Title Mismatched")
            test_check.writeResultExclReport(self, actualData=str(pageTitle_actual),
                                             dataToVerify=str(pageTitle_expected),
                                             result="Fail")



        Gateway_serialNumber = self.FrameworkInitializer.GetTestDataFromDataTable("Meter_Serial_Number")
        print(Gateway_serialNumber)

        serialnumberInputBox_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading","Meter Serial Number")
        print(serialnumberInputBox_locator)
        self.FrameworkCommonLibrary.Driver.find_element_by_id(serialnumberInputBox_locator[1]).send_keys(
            Gateway_serialNumber)


        submitBtn_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading", "Submit_button")
        self.FrameworkCommonLibrary.Driver.find_element_by_id(submitBtn_locator[1]).click()


        time.sleep(10)
        """message_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading", "Expected Message")
        messageReceived = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(message_locator[1]).text
        expectedMessage = self.FrameworkInitializer.GetTestDataFromDataTable("Expected Message")

        if str(messageReceived) == str(expectedMessage):
            print("Message checked and passed")
            test_check.writeResultExclReport(self, actualData=str(messageReceived),
                                             dataToVerify=str(expectedMessage),
                                             result="Pass")
        else:
            print("Message Checking Failed")
            test_check.writeResultExclReport(self, actualData=str(messageReceived),
                                             dataToVerify=str(expectedMessage),
                                             result="Fail")
        time.sleep(2)

        # press ok button
        ObBtn_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading", "OkBtn")
        self.FrameworkCommonLibrary.Driver.find_element_by_xpath(ObBtn_locator[1]).click()
        time.sleep(4)

        # sleep for 8 minutes
        time.sleep(480)
        # page refrese
        self.FrameworkCommonLibrary.Driver.refresh()
        time.sleep(10)"""


        import json
        substring = self.FrameworkInitializer.GetTestDataFromDataTable("Replaceable String")
        substring = json.loads(substring)
        key_list = list(substring.keys())
        val_list = list(substring.values())

        Query = self.FrameworkInitializer.GetTestDataFromDataTable("Sql Query")
        fullstring = str(Query)

        for x in range(len(key_list)):
            if key_list[x] in fullstring:
                print("Found!")
                fullstring = fullstring.replace(key_list[x], val_list[x])
            else:
                print("Not Found")
        print(fullstring)

        connectionString = self.FrameworkInitializer.GetTestDataFromDataTable("Connection String")
        sqlQuery = fullstring
        import pyodbc
        conn = pyodbc.connect(connectionString)
        cursor = conn.cursor()
        print("sqlQuery is = " + sqlQuery)
        cursor.execute(sqlQuery)
        conn.commit()

        # ***************
        # print("Current Iteration is = " +str(self.ObjSeleniumConfiguration.intCurrentIteration))
        self.ObjSeleniumConfiguration.intCurrentIteration = self.ObjSeleniumConfiguration.intCurrentIteration + 1
        counter = 0
        rowDataFetched = []
        serialNumber_database = ''
        #status_database = ''
        cumulativeEnergy_database = ''
        readingDate_database = ''
        lastBillingEnergy_database = ''
        lastBillingDemand_database = ''
        lastBillingDate_database = ''
        powerfactor_database = ''
        try:
            for row in cursor:
                counter = counter + 1
                print("[Info]: Database row imported = " + str(row))
                rowDataFetched.append(str(row))
                serialNumber_database = row[0]
                #status_database = ''
                cumulativeEnergy_database = row[1]
                readingDate_database = row[2]
                lastBillingEnergy_database = row[3]
                lastBillingDemand_database = row[4]
                lastBillingDate_database = row[5]
                powerfactor_database = row[6]
        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
        print(rowDataFetched[0])
        """serialNumber_database = ''
        status_database = ''
        cumulativeEnergy_database = ''
        readingDate_database = ''
        lastBillingEnergy_database = ''
        lastBillingDemand_database = ''
        lastBillingDate_database = ''
        powerfactor_database = ''"""

        databaseValues = [serialNumber_database, cumulativeEnergy_database, readingDate_database,
                          lastBillingEnergy_database
            , lastBillingDemand_database, lastBillingDate_database, powerfactor_database]
        print("dB VALUES")
        print(databaseValues)
        databaseValuesText = ["serialNumber", "cumulativeEnergy", "readingDate", "lastBillingEnergy",
                              "lastBillingDemand",
                              "lastBillingDate", "powerfactor"]

        serialNumber_web = ''
        cumulativeEnergy_database_web = ''
        readingDate_database_web = ''
        lastBillingEnergy_database_web = ''
        lastBillingDemand_database_web = ''
        lastBillingDate_database_web = ''
        powerfactor_database_web = ''

        try:
            serialNumber_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading",
                                                                        "Serial_Number")
            serialNumber_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(serialNumber_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)



        """status_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading", "Status")
        status_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(status_locator[1]).text"""

        try:
            cumulativeEnergy_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading",
                                                                            "Cummulative_Energy")
            cumulativeEnergy_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                cumulativeEnergy_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)


        try:
            readingDate_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading",
                                                                       "Reading_Date_Time")
            readingDate_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                readingDate_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)


        try:
            pass
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)



        try:
            lastBillingEnergy_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading",
                                                                             "Last_billing_Energy")
            lastBillingEnergy_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                lastBillingEnergy_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)


        try:
            lastBillingDemand_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading",
                                                                             "Last_Billing_Demand")
            lastBillingDemand_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                lastBillingDemand_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)


        try:
            lastBillingDate_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading",
                                                                           "Last_Billing_Date_Time")
            lastBillingDate_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                lastBillingDate_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)


        try:
            powerfactor_locator = self.FrameworkCommonLibrary.Locators("View_Meter_Reading_Page", "View_Reading",
                                                                       "Last_Billing_Power_Factor")
            powerfactor_database_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(
                powerfactor_locator[1]).text
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)


        webValues = [serialNumber_web, cumulativeEnergy_database_web, readingDate_database_web,
                     lastBillingEnergy_database_web,
                     lastBillingDemand_database_web, lastBillingDate_database_web, powerfactor_database_web]

        for x in range(len(databaseValues)):
            if str(databaseValues[x]) == str(webValues[x]):
                print(str(databaseValuesText[x]) + " : value matched = " + str(webValues[x]))
                test_check.writeResultExclReport(self,
                                                 actualData=str(databaseValuesText[x]) + " : " + str(webValues[x]),
                                                 dataToVerify=str(databaseValuesText[x]) + " : " + str(
                                                     databaseValues[x]),
                                                 result="Pass")
            else:
                print(str(databaseValuesText[x]) + " : value Mismatched = " + str(webValues[x]))
                test_check.writeResultExclReport(self,
                                                 actualData=str(databaseValuesText[x]) + " : " + str(webValues[x]),
                                                 dataToVerify=str(databaseValuesText[x]) + " : " + str(
                                                     databaseValues[x]),
                                                 result="Fail")


    def GatewayDetailPageVerification(self):
        print("Gateway DetailPage Verification Started")

        #get the values from website
        totalGatewayCount_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                              "totalGatewayCount")
        serailNumber_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                              "serialNumber")
        simConnectionStatus_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                              "simConnectionStatus")
        gatewayDisconnectionReason_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                              "gatewayConnectionStatus")
        lastConnected_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                              "lastConnectedStatus")
        firmwareVersion_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                              "firmwareVersion")
        metersRead_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                              "meterReadByGateway")
        averageMeterRead_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                              "averageMeterReadGateway")

        serviceprovide_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                                        "serviceProvider")

        accessTechnology_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                                        "accessTechnology")
        meterDiscovered_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                                        "meterDiscovered")
        lastSync_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                                       "lastSync")
        lastReboot_locator = self.FrameworkCommonLibrary.Locators("Gateway_Detailed_Page", "gatewayDetails",
                                                                       "lastReboot")
       
       
        totalGatewayCount_web = self.FrameworkCommonLibrary.Driver.find_element_by_xpath(totalGatewayCount_locator[1]).text
        serailNumber_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(serailNumber_locator[1]).text
        simConnectionStatus_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(simConnectionStatus_locator[1]).text
        gatewayDisconnectionReason_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(gatewayDisconnectionReason_locator[1]).text
        lastConnected_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(lastConnected_locator[1]).text
        firmwareVersion_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(firmwareVersion_locator[1]).text
        metersRead_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(metersRead_locator[1]).text
        averageMeterRead_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(averageMeterRead_locator[1]).text
        serviceProvide_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(serviceprovide_locator[1]).text
        accessTechnology_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(accessTechnology_locator[1]).text
        meterDiscovered_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(meterDiscovered_locator[1]).text
        lastSync_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(lastSync_locator[1]).text
        lastReboot_web = self.FrameworkCommonLibrary.Driver.find_element_by_id(lastReboot_locator[1]).text

        # read the values from the excel
        connectionString_1 = self.FrameworkInitializer.GetTestDataFromDataTable("Connection_String_1")
        connectionString_2 = self.FrameworkInitializer.GetTestDataFromDataTable("Connection_String_2")
        query1 = self.FrameworkInitializer.GetTestDataFromDataTable("Query_1")
        query2= self.FrameworkInitializer.GetTestDataFromDataTable("Query_2")
        replaceString_1 = self.FrameworkInitializer.GetTestDataFromDataTable("ReplaceString_1")
        replaceString_2 = self.FrameworkInitializer.GetTestDataFromDataTable("ReplaceString_2")

        #Executing Query 1*********************************************************************************************

        import json
        substring = replaceString_1
        substring = json.loads(substring)
        key_list = list(substring.keys())
        val_list = list(substring.values())

        Query = query1
        fullstring = str(Query)


        for x in range(len(key_list)):
            if key_list[x] in fullstring:
                print("Found!")
                fullstring = fullstring.replace(key_list[x], val_list[x])
            else:
                print("Not Found")
    #print(fullstring)

        sqlQuery = fullstring
        try:
            import pyodbc
            conn = pyodbc.connect(connectionString_1)
            cursor = conn.cursor()
            # print("sqlQuery is = " + sqlQuery)
            cursor.execute(sqlQuery)
            conn.commit()
        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        # print("Current Iteration is = " +str(self.ObjSeleniumConfiguration.intCurrentIteration))
        self.ObjSeleniumConfiguration.intCurrentIteration = self.ObjSeleniumConfiguration.intCurrentIteration + 1
        counter = 0
        rowDataFetched = []
        serialNumber_database = ''
        communicationStatus_database = ''
        statusReason_database = ''
        firmwareVersion_database = ''
        totalGatewayCount_database = ''
        lastConnectedSince_database = ''

        try:
            for row in cursor:
                counter = counter + 1
                print("[Info]: Database row imported = " + str(row))
                rowDataFetched.append(str(row))
                serialNumber_database = row[0]
                communicationStatus_database = row[3]
                statusReason_database = row[4]
                firmwareVersion_database = row[5]
                totalGatewayCount_database = row[6]
                lastConnectedSince_database = row[2]

        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)


        # Executing Query 2****************************************************************************

        import json
        substring2 = replaceString_2
        substring2 = json.loads(substring2)
        key_list2 = list(substring2.keys())
        val_list2 = list(substring2.values())

        Query2 = query2
        fullstring2 = str(Query2)

        for x in range(len(key_list2)):
            if key_list2[x] in fullstring2:
                print("Found!")
                fullstring2 = fullstring2.replace(key_list2[x], val_list2[x])
            else:
                print("Not Found")
        # print(fullstring)
        sqlQuery2 = fullstring2
        try:
            import pyodbc
            conn2 = pyodbc.connect(connectionString_2)
            cursor2 = conn2.cursor()
            # print("sqlQuery is = " + sqlQuery)
            cursor2.execute(sqlQuery2)
            conn2.commit()
        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        # print("Current Iteration is = " +str(self.ObjSeleniumConfiguration.intCurrentIteration))
        self.ObjSeleniumConfiguration.intCurrentIteration = self.ObjSeleniumConfiguration.intCurrentIteration + 1
        counter2 = 0
        rowDataFetched2 = []
        signalBar_database = ''
        serviceProvider_database = ''
        lastReboot_database = ''
        lastSync_database = ''
        meterDiscovery_database = ''
        todayCount_database = ''
        averageCount_database = ''
        accessTechnology_database = ''




        try:
            for row2 in cursor2:
                counter2 = counter2 + 1
                print("[Info]: Database row imported = " + str(row2))
                rowDataFetched2.append(str(row2))
                signalBar_database = row2[0]
                serviceProvider_database = row2[1]
                accessTechnology_database = row2[2]
                lastReboot_database = row2[3]
                lastSync_database = row2[4]
                meterDiscovery_database = row2[5]
                todayCount_database = row2[6]
                averageCount_database = row2[7]
        except Exception as e:
            print(str(e))
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        databaseValues = [serialNumber_database,communicationStatus_database,
                          statusReason_database,firmwareVersion_database,
                          totalGatewayCount_database,
                          serviceProvider_database,accessTechnology_database,
                          meterDiscovery_database,todayCount_database,averageCount_database,lastReboot_database,lastSync_database,
                          lastConnectedSince_database]

        databaseValuesText = ["serialNumber_database", "communicationStatus_database",
                          "statusReason_database", "firmwareVersion_database",
                          "totalGatewayCount_database",
                          "serviceProvider_database","accessTechnology_database",
                          "meterDiscovery_database", "todayCount_database", "averageCount_database","lastReboot_database","lastSync_database",
                              "lastConnectedSince_database"]



        webValues = [serailNumber_web,simConnectionStatus_web,gatewayDisconnectionReason_web,firmwareVersion_web,
                     totalGatewayCount_web,serviceProvide_web,accessTechnology_web,meterDiscovered_web,metersRead_web,averageMeterRead_web,lastReboot_web,
                     lastSync_web,lastConnected_web]




        for x in range(len(databaseValues)):
            if str(databaseValues[x]) == str(webValues[x]):
                print(str(databaseValuesText[x]) + " : value matched = " + str(webValues[x]))
                test_check.writeResultExclReport(self,
                                                 actualData=str(databaseValuesText[x]) + " : " + str(webValues[x]),
                                                 dataToVerify=str(databaseValuesText[x]) + " : " + str(
                                                     databaseValues[x]),
                                                 result="Pass")
            else:
                print(str(databaseValuesText[x]) + " : value Mismatched = " + str(webValues[x]))
                test_check.writeResultExclReport(self,
                                                 actualData=str(databaseValuesText[x]) + " : " + str(webValues[x]),
                                                 dataToVerify=str(databaseValuesText[x]) + " : " + str(
                                                     databaseValues[x]),
                                                 result="Fail")

















