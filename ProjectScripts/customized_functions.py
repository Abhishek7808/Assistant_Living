"""
 Version = 1.0.0.0
 -------------------------------------------------------------------------------
 Name: customized_functions.py
 Purpose: To define the customized functions to be used for keyword sheets.
 Author: Vivek Purohit (44454)
 -------------------------------------------------------------------------------
"""
import glob
import os
import stat
import sys
import time
# from time import sleep
from pythonframework.FrameworkCore.KeywordEngine.ReadKeywordTemplate import ReadKeywordTemplate as RKwT_Web

import pandas as pd
import self
from selenium.webdriver.common import alert
import pandas as pd
import openpyxl as xl
import psycopg2
import openpyxl as xl
import xlsxwriter
from selenium.webdriver.common.alert import Alert
from openpyxl import load_workbook

# from pythonframework.ProjectTemplate.CommonMethodLibrary.Appium_Android import Appium_Android
# from pythonframework.ProjectTemplate.CommonMethodLibrary.SeleniumCommonLibrary import SeleniumCommonLibrary


import pandas as pd
from pythonframework.FrameworkCore.ReadExcelData.ReadExcel import ReadExcel
from pythonframework.FrameworkCore.ObjectRepository.ObjectRepository import TestObject, SubObject
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import ConfigFile
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import Control_TestCase_File
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import AutomationApproach
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import ORFormat
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import ORApplicable
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import ScriptlessOR
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import ScriptlessORKeywords
from pythonframework.FrameworkCore.SeleniumConfiguration.Constants import TestCaseResults
from pythonframework.FrameworkCore.FrameworkLibrary.FrameworkCommonLibrary import FrameworkCommonLibrary
from pythonframework.FrameworkCore.CoreLibrary.SeleniumCoreLibrary import SeleniumCoreLibrary
from pythonframework.FrameworkCore.CoreLibrary.AppiumCoreLibrary import AppiumCoreLibrary
from pythonframework.FrameworkCore.TestResultReport.ColorPrint import ColorPrint
from pythonframework.FrameworkCore.SeleniumConfiguration.ConfigurationManager import ConfigurationManager
import time
from datetime import datetime, timedelta
import subprocess
from selenium.webdriver.common.by import By

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from appium.webdriver.common.touch_action import TouchAction
import paho.mqtt.client as mqtt
import json
import ssl
import paramiko

class test_check:

    def __init__(self, FrameworkInitializer):
        self.frameworkInitializer = FrameworkInitializer
        self.Driver = FrameworkInitializer.Driver
        self.AppiumDriver = FrameworkInitializer.AppiumDriver
        self.AppiumAndroidDriver = FrameworkInitializer.AppiumAndroidDriver
        self.AppiumIosDriver = FrameworkInitializer.AppiumIosDriver
        self.AppiumServer = None
        self.ObiterTestResult = FrameworkInitializer.ObiterTestResult
        self.FrameworkInitializer = FrameworkInitializer
        self.ObiterSeleniumConfiguration = FrameworkInitializer.ObiterSeleniumConfiguration
        self.FrameworkCommonLibrary = FrameworkInitializer.ObiterFrameworkCommon
        self.ObiterConfigurationManager = FrameworkInitializer.ObiterConfigurationManager
        self.Appium_Android = Appium_Android(FrameworkInitializer)
        self.ObiterSeleniumCommonLib = SeleniumCommonLibrary(FrameworkInitializer)
        self.ObjectRepository = FrameworkInitializer.ObjectRepository


custom_result_data = []


def execute_role_management(driver, data_iteration_from, data_iteration_to):
    """
    Function to perform automation for reusable component of
    iam --> role management.
    :param driver: Instantiate the driver.
    :return:
    """
    df_role_management = pd.read_excel(r"..\SeleniumAutomation\DataBase\Role_Management_Dataset.xlsx")
    for data_row in range(data_iteration_from, data_iteration_to):

        time.sleep(4)
        driver.get("https://insightval.ewatch.online/UserManagement/#/roles/add;fromList=1'")
        time.sleep(5)

        # iterate for each checkbox element
        module_name = df_role_management['Modules'].values[data_row]
        print(module_name)
        configuration_name_xpath = "//*[@class='UM_moduletreelistrepet']//span[contains(text(),'" \
                                   + module_name \
                                   + "')]"
        if driver.find_element_by_xpath(configuration_name_xpath).is_displayed() is True:
            driver.find_element_by_xpath(configuration_name_xpath).click()
            time.sleep(2)

        # get the number of rows in div
        configuration_row_count_xpath = "//*[@class='maindiv rightmain']"
        configuration_row_count = len(driver.find_elements_by_xpath(configuration_row_count_xpath))
        print(configuration_row_count)

        # get the row number in which the given configuration exists
        for row in range(1, configuration_row_count + 1):
            # create xpath for the row and find if the name of configuration settings exist
            configuration_setting_name = df_role_management['Configuration'].values[data_row]

            configuration_setting_xpath = "//*[@class='maindiv rightmain'][" \
                                          + str(row) \
                                          + "]//span[contains(text(),'" \
                                          + configuration_setting_name \
                                          + "')]"

            # check if element is present
            try:
                configuration_element = driver.find_element_by_xpath(configuration_setting_xpath)
                if configuration_element.is_displayed() is True:
                    row_number = int(row)
                    print(configuration_element.text)
                    print(row_number)

                    checkbox_permission = df_role_management[
                        'Permissions (Comma Seperated)'].values[data_row].split(",")

                    print(checkbox_permission)
                    print(len(checkbox_permission))

                    # Iterate for number of checkbox
                    for i in range(1, len(checkbox_permission) + 1):
                        checkbox_xpath = "(//*[@class='maindiv rightmain'][" \
                                         + str(row_number) + "]//span[@class='checkmark_check'])[" \
                                         + str(i) \
                                         + "]"
                        print(checkbox_xpath)

                        if checkbox_permission[i - 1].upper() == 'YES':
                            driver.find_element_by_xpath(checkbox_xpath).click()

            except Exception as err:  # pylint: disable=broad-except
                print(str(err))


def compareViewerAndExport(driver, data_iteration_from, data_iteration_to):
    import pandas as pd
    data1 = {"country": ["India", "USA", "UK", "Germany"], "dial_code": [91, 1, 44, 49]}
    df1 = pd.DataFrame(data1)
    df1.to_csv(r"C:\Users\44454\Desktop\data1.csv", index=None)
    data2 = {"country": ["India", "USA", "UK", "Germany", "Australia", "China"], "dial_code": [91, 1, 44, 49, 61, 86]}
    df2 = pd.DataFrame(data2)
    df2.to_csv(r"C:\Users\44454\Desktop\data2.csv", index=None)
    print(df1, "\n")
    print(df2)
    df1 = pd.read_csv(r"C:\Users\44454\Desktop\data1.csv")
    df2 = pd.read_csv(r"C:\Users\44454\Desktop\data2.csv")
    print(df1, "\n")
    print(df2, "\n")
    c_result = df1[df1.apply(tuple, 1).isin(df2.apply(tuple, 1))]
    print(c_result)
    custom_result = 'To valida', 'TestID', 'Pass', 'dataValue', 'var', 'dt_string'
    # custom_result_data.append(cust_result)


def sugam_billing_cash_counter(driver, data_iteration_from, data_iteration_to):
    """
    to calculate the number of notes to be fed to the system
    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :return:
    """
    try:
        print("inside custom function sugam_billing_cash_counter")
        element = driver.find_element_by_id("txtCollectionAmount")
        total_amount = element.get_attribute('value')
        total_amount = int(float(total_amount))
        print(f"total amount found = {total_amount}")

        two_thousand_notes = int(int(total_amount) / 2000)
        print(f"two_thousand_notes= {two_thousand_notes}")
        one_rupee_notes = int(int(total_amount) - int(two_thousand_notes * 2000))
        print(f"one_rupee_notes = {one_rupee_notes}")

        driver.get("https://10.10.103.222/0BSW/BillingPortal//GeneratePayInSlip/PayInSlip")

        driver.find_element_by_id("RupeestextboxTEXT2000").send_keys(two_thousand_notes)
        driver.find_element_by_id("RupeestextboxTEXT1").send_keys(one_rupee_notes)
    except Exception as e:
        print(str(e))
        exc_type, exc_obiter, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)


def alert_popup_ok(driver, data_iteration_from, data_iteration_to):
    alert.accept()


def database(driver, data_iteration_from, data_iteration_to):
    '''Database access '''

    excel_file = 'D:\\Python Framework Keyword\\embc\\Insight\\SeleniumAutomation\\DataBase\\data_excel.xlsx'
    df = pd.read_excel(excel_file)

    excel_file_1 = 'D:\\Python Framework Keyword\\embc\\Insight\\SeleniumAutomation\\ControlFile\\Keyword\\KeywordControlFile.xlsx'
    df1 = pd.read_excel(excel_file_1)

    row_number = df1[df1['Test Scenario Name'] == 'Database'].index
    database_row_number = list(row_number)
    try:
        for i in database_row_number:
            if df1.at[i, 'Action'] == 1:
                data_iteration_from = df1.at[i, 'Data Iteration From']
                data_iteration_to = df1.at[i, 'Data Iteration To']

                for data_iteration_from in range(data_iteration_from - 1, data_iteration_to):
                    conn = psycopg2.connect(database=df.at[data_iteration_from, 'database'],
                                            host=df.at[data_iteration_from, 'host'],
                                            user=df.at[data_iteration_from, 'user'],
                                            port=df.at[data_iteration_from, 'port'],
                                            password=df.at[data_iteration_from, 'password'])
                    cursor = conn.cursor()
                    query = df.at[data_iteration_from, 'query']
                    data = pd.read_sql_query(query, conn)
                    pd.DataFrame.to_csv(data, df.at[data_iteration_from, 'path'], index=False)


    except Exception as e:
        print(str(e))
        exc_type, exc_obiter, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)


# Code to compare two csv

def compareCSV(driver, data_iteration_from, data_iteration_to):
    try:
        # code to get the last downloaded file and lst modified time
        folder_path = r'C:\Users\45035\Downloads'
        file_type = '\*csv'
        files = glob.glob(folder_path + file_type)
        max_file = max(files, key=os.path.getctime)
        print(max_file)

        # get the last modified time of the file
        fileStatsObiter = os.stat(max_file)
        modificationTime = time.ctime(fileStatsObiter[stat.ST_MTIME])
        print("Last Modified Time : ", modificationTime)

        # compare csv
        df1 = pd.read_csv(r"..\..\..\exported_data_excel.csv")
        df2 = pd.read_csv(max_file)
        df1.sort_values(df1.columns[0], axis=0, inplace=True)
        df2.sort_values(df2.columns[0], axis=0, inplace=True)
        print(df1)
        print(df2)

        # comparing headers
        if len(df1.columns) == len(df2.columns):
            if ((df1.axes[1] == df2.axes[1]).any()):  # ==> 1 is to identify columns

                print("Headers Matched")
                test_check.writeResultExclReport(self,
                                                 actualData="Headers Matched",
                                                 dataToVerify="Headers Matched",
                                                 result="Pass")

                if df1.empty and df2.empty:
                    print("Please check Empty values are identified in the exported Dataset")
                    test_check.writeResultExclReport(self,
                                                     actualData="Please check Empty values are identified in the exported Dataset - Name of file used for matching = " + str(
                                                         max_file),
                                                     dataToVerify="Last modified time of the file was = " + str(
                                                         modificationTime),
                                                     result="Pass")
                else:
                    comparison_values = df1.values == df2.values
                    print(type(comparison_values))
                    print(comparison_values)
                    # exists = (False == comparison_values)
                    exists = False in comparison_values
                    print(exists)
                    if str(exists).upper() == "TRUE":
                        ######################################Code to find the index where data matching failed###################
                        ''' Get index positions of value in dataframe i.e. dfObiter.'''
                        listOfPos = list()
                        # Get bool dataframe with True at positions where the given value exists
                        df = pd.DataFrame(comparison_values)
                        print(df)
                        result = df.isin([False])
                        print(result)
                        # Get list of columns that contains the value
                        seriesObiter = result.any()
                        columnNames = list(seriesObiter[seriesObiter == True].index)
                        # Iterate over list of columns and fetch the rows indexes where value exists
                        for col in columnNames:
                            rows = list(result[col][result[col] == True].index)
                            for row in rows:
                                listOfPos.append((row, col))

                            ##########################################################################################################
                        print("Data Matching Failed")
                        test_check.writeResultExclReport(self,
                                                         actualData="Name of file used for matching = " + str(
                                                             max_file) + "Index Location = " + str(listOfPos),
                                                         dataToVerify="Last modified time of the file was = " + str(
                                                             modificationTime),
                                                         result="Fail")

                    else:
                        print("Data Matching Passed")
                        test_check.writeResultExclReport(self,
                                                         actualData="Name of file used for matching = " + str(
                                                             max_file),
                                                         dataToVerify="Last modified time of the file was = " + str(
                                                             modificationTime),
                                                         result="Pass")

            else:
                print("Headers Mismatched")
                test_check.writeResultExclReport(self,
                                                 actualData="Headers Mismatched",
                                                 dataToVerify="Headers Mismatched",
                                                 result="Fail")
        else:
            print("Column Size Mismatched")
            test_check.writeResultExclReport(self,
                                             actualData="Headers of the compared CSV is Not Matching!!",
                                             dataToVerify="Headers of the compared CSV should Match",
                                             result="Fail")





    except Exception as e:
        print("Test Failed")
        print(str(e))
        test_check.writeResultExclReport(self, actualData="Error in matching file " + str(max_file),
                                         dataToVerify="Last modified time of the file was = " + str(
                                             modificationTime),
                                         result="Fail")
        exc_type, exc_obiter, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)


def custom_click(driver, data_iteration_from, data_iteration_to):
    try:
        driver.find_element_by_xpath("//*[@class='cancel-button']").click()
        alert = Alert(driver)
        alert_text = alert.text
        print(alert_text)
        alert.accept()
    except Exception as e:
        print(str(e))


def ConnectionDetails_verify(driver, data_iteration_from, data_iteration_to, framework_variables):
    global othernum2, SUMO_new_2, Sugam_new_2, endotherm3, consstat2, category2, f
    try:
        import pandas as pd
        import psycopg2
        import self
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        import time
        import re

        excel_file = 'D:\\PythonFrameworkKeyword\\embc\\Insight\\SeleniumAutomation\\DataBase\\Verify.xlsx'
        df4 = pd.read_excel(excel_file)
        database_row_iteration = (df4.at[0, 'Database row iteration'])
        da1 = list(database_row_iteration)
        iter_from = int(da1[1])
        iter_to = int(da1[3])
        print(da1)
        print(iter_from)
        print(iter_to)

        """Code to get full address of consumer from Sugam"""
        for data1 in range(iter_from - 1, iter_to):
            Cons_num = df4.at[data1, "SUMO Consumer number"]
            conn = psycopg2.connect(database="sugamdb",
                                    host='10.10.103.220',
                                    user='postgres',
                                    port='5432',
                                    password='postgres')
            cursor = conn.cursor()
            query = df4.at[data1, 'Address']
            print(query)
            query.replace("###", Cons_num)
            print(query)
            cursor.execute(query)
            address = cursor.fetchone()
            a = ''
            for item in address:
                a = a + item
                b = a.replace(" ", "")
                e = b.replace(",", "")
            f = e[:-14]
            print("Address on Sugam is: ", f)

            """Code to get name on Sugam"""
            query9 = df4.at[data1, 'Name']
            cursor.execute(query9)
            fullname = list(cursor.fetchone())
            sugam_name = str(fullname[0])
            print("Full name on Sugam is:", sugam_name)
            sugam_name1 = sugam_name.replace(",", "")
            sugam_name2 = sugam_name1.replace(" ", "")

            """Code to get sanctioned load from Sugam"""
            query1 = df4.at[data1, 'Sanctioned load']
            cursor.execute(query1)
            sanctionedload = cursor.fetchone()
            sl = list(sanctionedload)
            sl1 = sl[0]
            sl2 = float(sl1)
            print("Sanctioned load on Sugam is: ", sl2)

            """Code to get mobile number from Sugam"""
            query2 = df4.at[data1, 'Mobile number']
            cursor.execute(query2)
            mobilenumber = cursor.fetchone()
            if mobilenumber is None:
                mobilenumber = ''
                print("Mobile number on Sugam is: ", mobilenumber)
            else:
                mn = list(mobilenumber)
                mn1 = mn[0]
                mn2 = str(mn1)
                print("Mobile number on Sugam is:", mn2)

            """Code to get categoryid from Sugam"""
            query8 = df4.at[data1, 'Category']
            cursor.execute(query8)
            category = cursor.fetchone()
            category1 = list(category)
            category2 = category1[0]
            print("Category on Sugam is:", category2)
            if category2 == 64:
                category3 = 'Domestic'
            elif category2 == 65:
                category3 = 'Non-Domestic'
            else:
                category3 = 'Other'
            if category3 == 'Other':
                meterno3 = 'NO-METER'

            """Code to get email from Sugam"""

            query3 = df4.at[data1, 'Email']
            cursor.execute(query3)
            email = cursor.fetchone()
            email1 = list(email)
            email2 = email1[0]
            email3 = str(email2)
            if email3 is None:
                email3 = ''
                print("Email on Sugam is: ", email3)
            else:
                print("Email on Sugam is:", email3)

            """Code to get secondary mobile number"""
            query4 = df4.at[data1, 'Landline number']
            cursor.execute(query4)
            endotherm = cursor.fetchone()
            if endotherm is None:
                endotherm = ''
                print("Other number on Sugam is:", endotherm)
            else:
                endotherm1 = list(endotherm)
                endotherm2 = endotherm1[0]
                endotherm3 = str(endotherm2)
                print("Other number on Sugam is:", endotherm3)

            """Code to get connection status"""
            query7 = df4.at[data1, 'Connection status']
            cursor.execute(query7)
            consstat = cursor.fetchone()
            consstat1 = list(consstat)
            consstat2 = consstat1[0]
            if consstat2 == 1:
                connection = 'Active'
            elif consstat2 == 2:
                connection = 'Active'
            else:
                connection = 'In-active'
            print("Connection status on Sugam is:", connection)
            if connection == 'In-active':
                meterno3 = 'NO-METER'

            """Code to get meterno"""
            query5 = df4.at[data1, 'Meter number']
            cursor.execute(query5)
            meterno = cursor.fetchone()
            meterno1 = list(meterno)
            meterno2 = meterno1[0]
            meterno3 = str(meterno2)
            if meterno3 is None:
                meterno3 = ''
                print("Meter number on Sugam is:", meterno3)
            elif consstat2 == 3 or category2 != 64 and 65:
                print("Meter number on Sugam is:", meterno3)
                meterno3 = 'NO-METER'
            else:
                meterno3 = str(meterno2)

            """Code to get phase from Sugam"""
            query6 = df4.at[data1, 'Phase type']
            cursor.execute(query6)
            phasetype = cursor.fetchone()
            phasetype1 = list(phasetype)
            phasetype2 = phasetype1[0]
            phasetype3 = int(phasetype2)
            print("Phase type on Sugam is:", phasetype3)
            if phasetype3 == 1:
                meter_type_Sugam = 'Liberty 170 (5-60) - 1Ph'
                tariff = 'BANSCREDIT'
                Load_Limit = '18'
                Current_Limit = '75'
            else:
                meter_type_Sugam = 'Liberty 370 (10-100) - 3Ph'
                tariff = '370-CREDIT'
                Load_Limit = '90'
                Current_Limit = '125'
            print("Meter type on Sugam is:", meter_type_Sugam)
            print("Tariff on Sugam is:", tariff)
            print("Load limit on Sugam is:", Load_Limit)
            print("Current limit on Sugam is:", Current_Limit)

            """Code to get all values from SUMO"""
            """Address"""
            driver = webdriver.Chrome("C:\\Users\\45125\\Documents\\chromedriver.exe")
            driver.maximize_window()
            driver.get("https://demo.liberty.online/")
            driver.find_element("id", "LoginId").send_keys("1130045125")
            driver.find_element("id", "Password").send_keys("System@1234")
            driver.find_element("id", "btnLogin").click()
            driver.find_element("xpath", "(//*[@class='level1Menu'])[3]").click()
            time.sleep(5)
            driver.find_element(By.ID, "sidemenu_304").click()
            time.sleep(2)
            driver.find_element(By.XPATH, "//*[@class='dropdown-toggle searchinput']").click()
            time.sleep(2)
            driver.find_element("xpath", "//*[@class='dropdown-toggle searchinput']").send_keys(
                int(df4.at[data1, 'SUMO Consumer number']))
            driver.find_element("id", "btnSearch").click()
            time.sleep(5)

            house_no = driver.find_element("id", "txtHouseNo").get_attribute("value")
            address_line_1 = driver.find_element("id", "txtAddress1").get_attribute("value")
            address_line_2 = driver.find_element("id", "txtAddress2").get_attribute("value")
            address_line_3 = driver.find_element("id", "txtAddress3").get_attribute("value")
            address_line_4 = driver.find_element("id", "txtAddress4").get_attribute("value")
            pincode = driver.find_element("id", "txtAddress4").get_attribute("value")
            full_address = house_no + " " + address_line_1 + " " + address_line_2 + " " + address_line_3 + " " + address_line_4 + " " + pincode
            c = ''
            for item in full_address:
                c = c + item
                d = c.replace(" ", "")
            print("Address on SUMO is: ", d)

            if f == d:
                print("Address on Sugam and SUMO matched")
                custom_result = 'Address Verify', 'TestID', 'Pass', f, d, 'dt_string'
                custom_result_data.append(custom_result)
            else:
                print("Address on Sugam and SUMO did not matched")
                custom_result = 'Address Verify', 'TestID', 'Fail', f, d, 'dt_string'
                custom_result_data.append(custom_result)

            """Name"""
            SUMO_First_name = driver.find_element("id", "txtFirstName").get_attribute("value")
            SUMO_Last_name = driver.find_element("id", "txtLastName").get_attribute("value")
            SUMO_full_name = SUMO_First_name + SUMO_Last_name
            print("Full name on SUMO is:", SUMO_full_name)

            SUMO_new = ''
            for item in SUMO_full_name.replace("None", ""):
                SUMO_new = SUMO_new + item
                SUMO_new_1 = SUMO_new.replace(" ", "")
                SUMO_new_2 = SUMO_new_1.replace(",", "")

            if str(sugam_name2) == str(SUMO_new_2):
                print("Name on Sugam and SUMO matched")
                custom_result = 'Name Verify', 'TestID', 'Pass', sugam_name2, SUMO_new_2, 'dt_string'
                custom_result_data.append(custom_result)

            else:
                print("Name on Sugam and SUMO did not matched")
                custom_result = 'Name Verify', 'TestID', 'Fail', sugam_name2, SUMO_new_2, 'dt_string'
                custom_result_data.append(custom_result)

            "Sanctioned load"
            SL_SUMO = driver.find_element("id", "txtSanctionedLoad").get_attribute("value")
            print("Sanctioned load on SUMO is:", SL_SUMO)
            if float(sl2) == float(SL_SUMO):
                print("Sanctioned load on Sugam and SUMO are equal")
                custom_result = 'Sanctioned load Verify', 'TestID', 'Pass', sl2, SL_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            else:
                print("Sanctioned load on Sugam and SUMO are not equal")
                custom_result = 'Sanctioned load Verify', 'TestID', 'Fail', sl2, SL_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            """Mobile number"""

            Mobile_SUMO = driver.find_element("id", "txtPhone1").get_attribute("value")
            print("Mobile number of SUMO is:", Mobile_SUMO)
            if Mobile_SUMO == 'None':
                Mobile_SUMO = int()
            if str(mn2) == str(Mobile_SUMO):
                print("Mobile number on Sugam and SUMO are equal")
                custom_result = 'Mobile number Verify', 'TestID', 'Pass', mn2, Mobile_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            else:
                print("Mobile number on Sugam and SUMO are not equal")
                custom_result = 'Mobile number Verify', 'TestID', 'Fail', mn2, Mobile_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            """Secondary mobile number"""
            Mobile_SEC_SUMO = driver.find_element("id", "txtPhone2").get_attribute("value")
            print("Other number on SUMO is:", Mobile_SEC_SUMO)
            if endotherm3 == Mobile_SEC_SUMO:
                print("Secondary mobile number on Sugam and SUMO are equal")
                custom_result = 'Landline Verify', 'TestID', 'Pass', endotherm3, Mobile_SEC_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            else:
                print("Secondary mobile number on Sugam and SUMO are not equal")
                custom_result = 'Landline Verify', 'TestID', 'Fail', endotherm3, Mobile_SEC_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            """Code to get type of meter"""
            metertype_SUMO = driver.find_element("xpath", "(//*[@class='col-lg-5 col-md-6 col-sm-7 col-xs-6'])[3]").text
            print("Meter type on SUMO is:", metertype_SUMO)
            if meter_type_Sugam == metertype_SUMO:
                print("Meter type on Sugam and SUMO are equal")
                custom_result = 'Meter type Verify', 'TestID', 'Pass', meter_type_Sugam, metertype_SUMO, 'dt_string'
                custom_result_data.append(custom_result)
            else:
                print("Meter type on Sugam and SUMO are not equal")
                custom_result = 'Meter type Verify', 'TestID', 'Fail', meter_type_Sugam, metertype_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            """Code to get connection status"""
            constatus_SUMO = driver.find_element("xpath", "(//*[@class='col-lg-5 col-md-6 col-sm-7 col-xs-6'])[7]").text
            print("Connection status on SUMO is", constatus_SUMO)
            if connection == constatus_SUMO and category3 != 'Other':
                print("Connection status on Sugam and SUMO matched")
                custom_result = 'Connection Status Verify', 'TestID', 'Pass', connection, constatus_SUMO, 'dt_string'
                custom_result_data.append(custom_result)
            else:
                print("Connection status on Sugam and SUMO did not matched")
                custom_result = 'Connection Status Verify', 'TestID', 'Fail', connection, constatus_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            """Code to verify meter number on SUMO"""
            meter_SUMO = driver.find_element("xpath", "(//*[@class='col-lg-5 col-md-6 col-sm-7 col-xs-6'])[4]").text
            print("Meter no on SUMO is", meter_SUMO)
            if meterno3 == meter_SUMO:
                print("Meter number on Sugam and SUMO matched")
                custom_result = 'Meter Numbner Verify', 'TestID', 'Pass', meterno3, meter_SUMO, 'dt_string'
                custom_result_data.append(custom_result)
            else:
                print("Meter number on Sugam and SUMO did not matched")
                custom_result = 'Meter Numbner Verify', 'TestID', 'Fail', meterno3, meter_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            """Code to verify tariff code on SUMO"""
            tariff_SUMO = driver.find_element("xpath", "//*[@class='dropdownbutton-margin']").text
            print("Tariff on SUMO is", tariff_SUMO)
            if tariff == tariff_SUMO:
                print("Tariff on Sugam and SUMO matched")
                custom_result = 'Tariff code Verify', 'TestID', 'Pass', tariff, tariff_SUMO, 'dt_string'
                custom_result_data.append(custom_result)
            else:
                print("Tariff on Sugam and SUMO did not matched")
                custom_result = 'Tariff code Verify', 'TestID', 'Fail', tariff, tariff_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            """Code to verify load limit"""
            loadlimit_SUMO = driver.find_element("id", "txtLoadLimit").get_attribute("value")
            print("Load limit on SUMO is", loadlimit_SUMO)
            if Load_Limit == loadlimit_SUMO:
                print("Load limit on Sugam and SUMO matched")
                custom_result = 'Load limit Verify', 'TestID', 'Pass', Load_Limit, loadlimit_SUMO, 'dt_string'
                custom_result_data.append(custom_result)
            else:
                print("Load limit on Sugam and SUMO did not matched")
                custom_result = 'Load limit Verify', 'TestID', 'Fail', Load_Limit, loadlimit_SUMO, 'dt_string'
                custom_result_data.append(custom_result)

            """Code to verify current limit"""
            Currentlimit_SUMO = driver.find_element("id", "txtCurrentLimit").get_attribute("value")
            print("Current limit on SUMO is", Currentlimit_SUMO)
            if Current_Limit == Currentlimit_SUMO:
                print("Current limit on Sugam and SUMO matched")
                custom_result = 'Current limit Verify', 'TestID', 'Pass', Current_Limit, Currentlimit_SUMO, 'dt_string'
                custom_result_data.append(custom_result)
            else:
                print("Current limit on Sugam and SUMO did not matched")
                custom_result = 'Current limit Verify', 'TestID', 'Fail', Current_Limit, Currentlimit_SUMO, 'dt_string'
                custom_result_data.append(custom_result)


    except Exception as e:
        print("Test Failed")
        print(str(e))
        exc_type, exc_obiter, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)


def name_verify(driver, data_iteration_from, data_iteration_to, framework_variables):
    """

    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :param framework_variables:
    :return:
    """

    try:
        import pandas as pd
        import psycopg2
        import self
        import time

        from selenium import webdriver
        from selenium.webdriver.common.by import By

        """Code to get full name of consumer from Sugam"""

        conn = psycopg2.connect(database="sugamdb",
                                host='10.11.2.155',
                                user='postgres',
                                port='5432',
                                password='postgres')
        cursor = conn.cursor()
        query = "Select firstname, lastname, middlename, orgname from " \
                "cim.t_cim_cm_elecconndetails where connectionnumber = '130311020576'"
        cursor.execute(query)
        fullname = list(cursor.fetchone())
        first_name = fullname[0]
        middle_name = fullname[1]
        last_name = fullname[2]
        org_name = fullname[3]

        Sugam_full_name = str(first_name) + str(middle_name) + str(last_name) + str(org_name)
        print("Full name on Sugam is:", Sugam_full_name.replace("None", ""))
        Sugam_new = ''
        for item in Sugam_full_name.replace("None", ""):
            Sugam_new = Sugam_new + item
            Sugam_new_1 = Sugam_new.replace(" ", "")
            Sugam_new_2 = Sugam_new_1.replace(",", "")
        print(Sugam_new_2)

        """Code to get full name from SUMO"""

        driver = webdriver.Chrome("C:\\Users\\45125\\Documents\\chromedriver.exe")
        driver.maximize_window()
        driver.get("https://demo.liberty.online/")
        driver.find_element("id", "LoginId").send_keys("1130045125")
        driver.find_element("id", "Password").send_keys("System@1234")
        driver.find_element("id", "btnLogin").click()
        driver.find_element("xpath", "(//*[@class='level1Menu'])[3]").click()
        time.sleep(2)
        driver.find_element(By.ID, "sidemenu_304").click()
        time.sleep(2)
        driver.find_element(By.XPATH, "//*[@class='dropdown-toggle searchinput']").click()
        time.sleep(2)
        driver.find_element("xpath", "(//*[@class='searchlink'])[4]").click()
        time.sleep(2)
        driver.find_element("xpath", "//*[@class='dropdown-toggle searchinput']").send_keys("SS21499155")
        driver.find_element("id", "btnSearch").click()
        time.sleep(2)

        SUMO_First_name = driver.find_element("id", "txtFirstName").get_attribute("value")
        SUMO_Last_name = driver.find_element("id", "txtLastName").get_attribute("value")
        SUMO_full_name = SUMO_First_name + SUMO_Last_name
        print("Full name on SUMO is:", SUMO_full_name)

        SUMO_new = ''
        for item in Sugam_full_name.replace("None", ""):
            SUMO_new = SUMO_new + item
            SUMO_new_1 = SUMO_new.replace(" ", "")
            SUMO_new_2 = SUMO_new_1.replace(",", "")
        print(SUMO_new_2)

        if Sugam_new_2 == SUMO_new_2:
            print("Name on Sugam and SUMO are equal")
            print("Inside custom function")
            custom_result = 'Description', 'TestID', 'Pass', Sugam_new_2, SUMO_new_2, 'dt_string'
            custom_result_data.append(custom_result)

        else:
            print("Name on Sugam and SUMO are not equal")
            print("Inside custom function")
            custom_result = 'Description', 'TestID', 'Fail', Sugam_new_2, SUMO_new_2, 'dt_string'
            custom_result_data.append(custom_result)

    except Exception as e:
        print("Test Failed")
        print(str(e))
        exc_type, exc_obiter, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)


def demo_function(driver, data_iteration_from, data_iteration_to, framework_variables):
    print("This is a Demo Function and it is running Fine")


def sumo_web_check_tariff_open(driver, data_iteration_from, data_iteration_to, framework_variables):
    """
    This function checks that Weather the Configuration management>Tariff option
    is open or not on SUMO, If open it pass and If close it opens it
    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :param framework_variables:
    :return:
    """

    print("Customised Function Started")
    time.sleep(2)
    x = driver.find_element_by_id('sidemenu_201').is_displayed()  # look for Tariff category Option
    if x == 1:
        print('It\'s Open Now, No Problem')
    elif x == 0:
        print('It\'s not Open')
        driver.find_element_by_id('sidemenu_TARIFF').click()  # click on Configuration management
        time.sleep(1)
        sumo_web_check_tariff_open(driver, data_iteration_from, data_iteration_to, framework_variables)
    else:
        print('Else Condition, Please Check !!')


def sumo_web_check_vending_inner_open(driver, data_iteration_from, data_iteration_to, framework_variables):
    """
    This function checks that Weather Transactions>Vending page is open or not on SUMO,
    and it opens it
    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :param framework_variables:
    :return:
    """

    print("Customised Function Started")
    time.sleep(2)
    x = driver.find_element_by_id('sidemenu_702').is_displayed()

    if x == 1:
        print('It\'s Open Now, No Problem')

    elif x == 0:
        print('It\'s not Open')
        driver.find_element_by_id('sidemenu_TRANSACTIONS').click()  # Click on Vending(Outer)
        time.sleep(1)

    else:
        print('Else Condition, Please Check !!')


def sumo_web_check_meter_reading_open(driver, data_iteration_from, data_iteration_to, framework_variables):
    """
    This function checks that Weather the Meter reading
    is open or not on SUMO, If open it pass and If close it opens it
    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :param framework_variables:
    :return:
    """

    print("Customised Function Started")
    time.sleep(2)
    x = driver.find_element_by_id('sidemenu_407').is_displayed()  # look for On-demand reading

    if x == 1:
        print('It\'s Open Now, No Problem')

    elif x == 0:
        print('It\'s not Open')
        driver.find_element_by_xpath('/html/body/div/div[1]/div/ul/li[5]/a').click()  # click on Meter reading
        time.sleep(1)
        sumo_web_check_meter_reading_open(driver, data_iteration_from, data_iteration_to, framework_variables)

    else:
        print('Else Condition, Please Check !!')


def sumo_web_check_meter_profile_open(driver, data_iteration_from, data_iteration_to, framework_variables):
    """
    This function checks that Weather the Meter profile > Meter state information
    is open or not on SUMO, If open it pass and If close it opens it
    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :param framework_variables:
    :return:
    """

    print("Customised Function Started")
    time.sleep(2)
    x = driver.find_element_by_id('sidemenu_310').is_displayed()  # look for Meter state info Option

    if x == 1:
        print('It\'s Open Now, No Problem')

    elif x == 0:
        print('It\'s not Open')
        driver.find_element_by_id('sidemenu_METER').click()  # click on Meter profile
        time.sleep(1)
        sumo_web_check_meter_profile_open(driver, data_iteration_from, data_iteration_to, framework_variables)

    else:
        print('Else Condition, Please Check !!')


def sumo_web_check_debt_management_open(driver, data_iteration_from, data_iteration_to, framework_variables):
    """
    This function checks that Weather the Transactions > Debt management
    is open or not on SUMO, If open it pass and If close it opens it
    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :param framework_variables:
    :return:
    """

    print("Customised Function Started")
    time.sleep(2)
    x1 = driver.find_element_by_id('sidemenu_702').is_displayed()  # look for vending inner
    x2 = driver.find_element_by_id('sidemenu_305').is_displayed()  # look for debt management inner

    if x1 == 0 and x2 == 0:
        print('Transactions is not Open')
        driver.find_element_by_id('sidemenu_TRANSACTIONS').click()  # click on transactions
        time.sleep(1)
        sumo_web_check_debt_management_open(driver, data_iteration_from, data_iteration_to, framework_variables)

    elif x1 == 1 and x2 == 0:
        print('Debt management is not Open')
        driver.find_element_by_id('sidemenu_DEBT').click()  # click on debt management
        time.sleep(1)
        sumo_web_check_debt_management_open(driver, data_iteration_from, data_iteration_to, framework_variables)

    elif x1 == 0 and x2 == 1:
        print('It\'s Open, no issue')

    else:
        print('Else Condition, Please Check !!')


def sumo_web_check_installation_open(driver, data_iteration_from, data_iteration_to, framework_variables):
    """
    This function checks that Weather the Installation & decommission menu
    is open or not on SUMO, If open it pass and If close it opens it
    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :param framework_variables:
    :return:
    """

    print("Customised Function Started")
    time.sleep(2)
    x = driver.find_element_by_id('sidemenu_401').is_displayed()  # look for Installation category Option

    if x == 1:
        print('It\'s Open Now, No Problem')

    elif x == 0:
        print('It\'s not Open')
        driver.find_element_by_id('sidemenu_INSTALLATION').click()  # click on Installation & decommission
        time.sleep(1)
        sumo_web_check_installation_open(driver, data_iteration_from, data_iteration_to, framework_variables)

    else:
        print('Else Condition, Please Check !!')


def sahaj_android_pending_token_check(driver, data_iteration_from, data_iteration_to, framework_variables):
    """
    This function checks that you have pending recharge tokens or not,
    If you don't have any pending tokens, it just pass,
    but If you have pending token, then it removes the token and then enter amount and select billdesk
    as a payment method and proceeds the script
    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :param framework_variables:
    :return:
    """

    print("Customised Function Started")
    x = driver.find_element_by_id('android:id/button1').is_displayed()

    if x == 1:
        print('Pending Token Present')
        driver.find_element_by_id('android:id/button1').click()  # click check now
        time.sleep(2)
        driver.find_element_by_id('com.sahaj.liberty.online.demo:id/btn_cancel').click()  # click cancel
        time.sleep(1)
        driver.find_element_by_id('android:id/button1').click()  # click yes
        time.sleep(1)
        driver.find_element_by_id('android:id/button1').click()  # click ok
        time.sleep(1)
        driver.find_element_by_xpath('//android.widget.LinearLayout[@content-desc="Details"]').click()  # click details
        driver.find_element_by_id('com.sahaj.liberty.online.demo:id/btn_recharge').click()  # click recharge
        time.sleep(2)
        driver.find_element_by_id(
            'com.sahaj.liberty.online.demo:id/buy_token_header_layout').click()  # click purchase token
        amount = framework_variables['dataValue']
        print('I am entering the amount provided in Data Sheet, i.e. = ', amount)
        driver.find_element_by_id('com.sahaj.liberty.online.demo:id/input_amount').send_keys(amount)  # enter amount
        driver.find_element_by_id('com.sahaj.liberty.online.demo:id/btn_online_pay').click()  # click proceed
        print('I don\'t know which payment method u r using, I am selecting Billdesk')
        driver.find_element_by_xpath("//*[contains(@text,'Pay via Billdesk')]").click()  # select billdesk
        driver.find_element_by_id('com.sahaj.liberty.online.demo:id/btnPay').click()  # click proceed

    elif x == 0:
        print('No Pending Token')

    else:
        print('Else Condition, Please Check !!')


def click_here(driver, data_iteration_from, data_iteration_to, framework_variables):
    print('Cus Fun Started')
    x = driver.find_element_by_id('FileObject').is_displayed()
    print('X = ', x)


def sumo_web_check_wan_status(driver, data_iteration_from, data_iteration_to, framework_variables):
    """
    This function gets the attribute of the image(that shows the wan and no-wan status)
    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :param framework_variables:
    :return:
    """

    print("Test Customised Function Started")
    x = driver.find_element_by_id('imgMeterWanStatus').get_attribute('src')
    print('src = ', x)


def write_data_to_datasheet(data_to_write, driver, data_iteration_from, data_iteration_to, framework_variables):
    # from openpyxl import load_workbook
    # write data to sheet
    write_df = pd.read_excel(framework_variables['DatabaseSheet'])
    column_header_list = []
    for col in write_df.columns:
        column_header_list.append(col)
    # value=data_to_write
    # row = framework_variables['datarow'] + 2
    columns = column_header_list.index(framework_variables['dataset_column_header']) + 1
    print(columns)
    workbook = load_workbook(filename=framework_variables['DatabaseSheet'])
    sheet = workbook.active
    sheet.cell(row=framework_variables['datarow'] + 2, column=columns).value = data_to_write
    workbook.save(filename=framework_variables['DatabaseSheet'])


def test_now_function(driver, data_iteration_from, data_iteration_to, framework_variables):
    print("Now Test Customised Function Started")
    value = driver.find_element_by_xpath('//*[@id="endpoint"]/span').text
    print('Value = ', value)
    write_data_to_datasheet(value, driver, data_iteration_from, data_iteration_to, framework_variables)
    # write_in_xl = RKwT_Web()
    # write_in_xl.K_WriteText_Datasheet(value)


def test_cus_fun(driver, data_iteration_from, data_iteration_to, framework_variables):
    """
    For learning and Testing
    :param driver:
    :param data_iteration_from:
    :param data_iteration_to:
    :param framework_variables:
    :return:
    """

    print('# Cus Fun Started')
    print('# Driver = ', driver)
    print('# Data itration from = ', data_iteration_from)
    print('# Data itration to = ', data_iteration_to)
    print('# Framework var = ', framework_variables)
    print('# Data Value = ', framework_variables['dataValue'])
    print('# The End')
    x = framework_variables['dataValue']
    print('# id of x = ', id(x))
    print('x = ', x)

    if x == '44':
        print('# Success')

    else:
        print('# Fail')

def select_destination_link(driver,data_iteration_from, data_iteration_to,framework_variables):
    locator_values = framework_variables['dataValue']
    locator_values = locator_values.split(',')

    time.sleep(1)
    # checkin status of 3rd value
    status_2 = driver.find_element(By.ID, locator_values[len(locator_values) - 1]).is_displayed()
    # checking status of 2nd value
    status_1 = driver.find_element(By.ID, locator_values[len(locator_values) - 2]).is_displayed()

    if not status_2 and not status_1:
        driver.find_element(By.ID, locator_values[0]).click()
        time.sleep(1)
        select_destination_link(driver, data_iteration_from, data_iteration_to, framework_variables)

    elif not status_2 and status_1:
        driver.find_element(By.ID, locator_values[len(locator_values) - 2]).click()
        time.sleep(1)
        select_destination_link(driver, data_iteration_from, data_iteration_to, framework_variables)

    elif status_2 and status_1:
        driver.find_element(By.ID, locator_values[len(locator_values) - 1]).click()
        time.sleep(1)

    else:
        print('Else Condition Please Check !!')

def sahaj_android_clear_update_popup(driver,data_iteration_from, data_iteration_to,framework_variables):
    """
    This function checks that 'Update Version' popup appears or not
    If popup appers it clears it and if no popup is there then it pass
    """
    print("Customised Function Started")
    time.sleep(1)
    x = driver.find_element_by_id('android:id/alertTitle').is_displayed()  # check for popup

    if x == 1:
        print('Popup Present')
        time.sleep(1)
        driver.find_element_by_id('android:id/button3').click()  # click remind me later
    elif x == 0:
        print('Popup is not Present')
    else:
        print('Else Condition Please Check !!')


def manual_wait(driver,data_iteration_from, data_iteration_to,framework_variables):
    """
    This function waits till you press 'enter' key manually
    """
    while True:
        print("")
        ent = input("Press the Enter key after manual work is done")
        if ent == "":
            print("Thanks for Confirmation")
            break


def driver_kill(driver,data_iteration_from, data_iteration_to,framework_variables):
    """App kill at running time"""
    driver.quit()
    print('App Closed')

def sahaj_android_select_DropDownValue(driver,data_iteration_from, data_iteration_to,framework_variables):
    """
    This function selects the mentioned Drop Down Value at Home Page of Sahaj Android App
    Value is mentioned in Data Sheet and colomn name is 'DropDownValue'
    """
    print("Customised Function Started")
    location_id = 'android:id/text'
    drop_down_value = framework_variables['dataValue']
    print("Drop Down Value =", drop_down_value)
    print('Clicking on ID =', location_id + drop_down_value)
    time.sleep(1)
    driver.find_element_by_id('com.sahaj.liberty.online.demo:id/title_drop').click()  # Clicking on Drop Down
    time.sleep(1)
    driver.find_element_by_id(location_id + drop_down_value).click()  # Clicking on the mentioned drop down value


    # if len(locator_values) == 3:
    #     status = driver.find_element(
    #         By.XPATH, f'//a[@id={locator_values[0]}]/following-sibling::ul').get_attribute('aria-expanded')
    #     print(status)
    #
    # elif len(locator_values) == 2:
    #     pass
    #
    # else:
    #     print('Check for the datasheet input.')

    print(len(locator_values))
    print(locator_values[0], locator_values[1], locator_values[2])


def generate_seconds_based_on_input(driver, data_iteration_from, data_iteration_to, framework_variables):
    def generate_date_based_on_input(input_value):
        # Define the base date (1988-01-01)
        base_date = datetime(1988, 1, 1)

        # Calculate the current date in seconds since 1988
        current_date = datetime.now()
        seconds_since_1988 = (current_date - base_date).total_seconds()

        # Calculate the target date based on the input value
        target_date = base_date + timedelta(seconds=seconds_since_1988 + input_value * 24 * 60 * 60)

        return target_date, seconds_since_1988 + input_value * 24 * 60 * 60

    # Example usage:
    input_value = int(framework_variables['dataValue'])  # Change this to the desired input value

    target_date, seconds_after_1988 = generate_date_based_on_input(input_value)

    if input_value > 0:
        print(f"Date after {input_value} days: {target_date.strftime('%Y-%m-%d')}")
    elif input_value < 0:
        print(f"Date before {abs(input_value)} days: {target_date.strftime('%Y-%m-%d')}")
    else:
        print(f"Midnight today: {target_date.strftime('%Y-%m-%d %H:%M:%S')}")

    print(f"Value in seconds since 1988: {int(seconds_after_1988)}")

    # Enter text
    try:
        text_input = driver.child_window(auto_id=framework_variables['locatorValue'],
                                         found_index=int(framework_variables['index']))
        text_input.wait('exists enabled visible', timeout=10)
        text_input.set_focus()

        # Set text in the text input (use a string instead of an integer)
        text_input.type_keys(int(seconds_after_1988))
    except Exception as e:
        print(str(e))
        import traceback
        traceback.print_exc()


def time_set_mlp_tool_sec_from_1988(driver,data_iteration_from, data_iteration_to,framework_variables):
    """

    """
    # Reference Date
    start = datetime.strptime("1988-01-01 0:0:00", "%Y-%m-%d %H:%M:%S")
    # Today's Date
    end = datetime.strptime(str(framework_variables['dataValue']), "%Y-%m-%d %H:%M:%S")

    difference = end - start

    seconds_after_1988 = difference.total_seconds()
    print('difference in seconds is:', seconds_after_1988)

    # Enter text
    try:
        text_input = driver.child_window(auto_id=framework_variables['locatorValue'],
                                         found_index=int(framework_variables['index']))
        text_input.wait('exists enabled visible', timeout=10)
        text_input.set_focus()

        # Set text in the text input (use a string instead of an integer)
        text_input.type_keys(int(seconds_after_1988))
    except Exception as e:
        print(str(e))
        import traceback
        traceback.print_exc()

def AL_Demo_Function(driver, data_iteration_from, data_iteration_to, framework_variables):
    print("Assisted Living Demo Custome Function")
    #print(framework_variables)
    print(f"[Info]: Data value retrieved form database = {framework_variables['dataValue']}")
    """result = [framework_variables['TestScenario'],
              framework_variables['TestID'],
              framework_variables['row'],
              framework_variables['datarow'],
              framework_variables['result'],
              framework_variables['actualData'],
              framework_variables['dataToVerify'],
              framework_variables['execution_time']]"""

    result = [framework_variables['TestScenario'],
          framework_variables['TestID'],
          framework_variables['row'],
          framework_variables['datarow'],
          framework_variables['dataValue'],
          "framework_variables['actualData']",
          "framework_variables['dataToVerify']",
          "framework_variables['execution_time']"]
    custom_result_data.append(result)

def compare_alert_id_function(driver, data_iteration_from, data_iteration_to, framework_variables):
    print("Assisted Living Demo Custome Function")
    #print(framework_variables)
    print(f"[Info]: Data value retrieved form database = {framework_variables['dataValue']}")

    values = framework_variables['dataValue'].split(',')

    # Assign the values to separate variables
    old_alert = values[0]
    new_alert = values[1]
    alert_result = values[2]

    # Print the results
    print("old_alert:", old_alert)
    print("new_alert:", new_alert)
    print("alert_result:", alert_result)

    result = [framework_variables['TestScenario'],
          framework_variables['TestID'],
          framework_variables['row'],
          framework_variables['datarow'],
          alert_result,
          old_alert,
          new_alert,
          "framework_variables['execution_time']"]
    custom_result_data.append(result)

def Open_Carer_App(driver, data_iteration_from, data_iteration_to, framework_variables):
    print("Inside carer app function")
    batch_file_path = r'D:\Automation Framework\Tablet Automation\Assisted_Living\ProjectScripts\OpenCarerApp.bat'
    subprocess.run([batch_file_path], shell=True)


def Find_Alert_On_Support_portal(driver, data_iteration_from, data_iteration_to, framework_variables):
    # Locate all rows of the table, including header
    rows = driver.find_elements(By.XPATH, "//table[@id='tblAlerts']/tbody/tr")
    print("Number of Rows: ", len(rows))

    # Iterate over each row, skipping the first row (assuming it's the header)
    for i, row in enumerate(rows, start=1):
        try:
            # Skip the header row if it contains <th> elements instead of <td>
            if row.find_elements(By.XPATH, ".//th"):
                print(f"Skipping header row {i}")
                continue

            # Locate the InstallationId element within the current row
            installation_id_element = row.find_element(By.XPATH, ".//td[2]")

            # Extract the text from the InstallationId cell
            installation_id_text = installation_id_element.text.strip()
            print(f"Row {i} Installation ID: {installation_id_text}")

            # Compare the InstallationId with the dataValue from framework_variables
            if str(framework_variables['dataValue']) == installation_id_text:
                print(f"Match Found: {installation_id_text} in Row {i}")
                # Locate the 'Accept' button in the same row
                accept_button = row.find_element(By.XPATH, ".//td[9]/a[@id='btnAccept']")
                print("Accept button found, clicking...")
                accept_button.click()
                break
        except Exception as e:
            print(f"Error processing row {i}: {e}")

def wait_and_unlock(driver, data_iteration_from, data_iteration_to, framework_variables):
    if driver.is_locked():
        # Unlock the device
        print(f"[Info]:Inside If Condition")
        driver.unlock()

        # If needed, swipe up to fully unlock the screen (in case of swipe lock)
        action = TouchAction(driver)
        print(f"[Info]:Inside Action")
        action.press(x=360, y=1200).wait(1000).move_to(x=360, y=300).release().perform()


def wait_until_element_visible_click(driver, data_iteration_from, data_iteration_to, framework_variables):
    # This Function will accept ID locator and only ID value not package Name.
    print("Inside Wait Until Element Visible Function")
    locator_value = framework_variables['dataValue']
    print("Here Is the Location:"+locator_value)
    try:
        # Initialize WebDriverWait with a 300-second timeout
        wait = WebDriverWait(driver, 300)

        # Wait until the element is visible, using the XPATH from framework_variables
        element = wait.until(EC.visibility_of_element_located((By.ID, locator_value)))

        print("Element is visible!")

        # Perform actions on the element if needed, e.g., click or read text
        element.click()  # Example action

        print("Click on Element")

    except Exception as e:
        print(f"An error occurred: {e}")

def wait_until_element_visible(driver, data_iteration_from, data_iteration_to, framework_variables):
    # This Function will accept ID locator and only ID value not package Name.
    print("Inside Wait Until Element Visible Function")
    locator_value = framework_variables['dataValue']
    print("Locator Value: "+locator_value)
    try:
        # Initialize WebDriverWait with a 300-second timeout
        wait = WebDriverWait(driver, 300)

        # Wait until the element is visible, using the XPATH from framework_variables
        element = wait.until(EC.visibility_of_element_located((By.ID, locator_value)))

        print("Element is visible!")

    except Exception as e:
        print(f"An error occurred: {e}")


def Close_Alert_On_Support_portal(driver, data_iteration_from, data_iteration_to, framework_variables):
    # Locate all rows of the table, including header
    rows = driver.find_elements(By.XPATH, "//table[@id='tblAlerts']/tbody/tr")
    print("Number of Rows: ", len(rows))

    # Iterate over each row, skipping the first row (assuming it's the header)
    for i, row in enumerate(rows, start=1):
        try:
            # Skip the header row if it contains <th> elements instead of <td>
            if row.find_elements(By.XPATH, ".//th"):
                print(f"Skipping header row {i}")
                continue

            # Locate the InstallationId element within the current row
            installation_id_element = row.find_element(By.XPATH, ".//td[2]")

            # Extract the text from the InstallationId cell
            installation_id_text = installation_id_element.text.strip()
            print(f"Row {i} Installation ID: {installation_id_text}")

            # Compare the InstallationId with the dataValue from framework_variables
            if str(framework_variables['dataValue']) == installation_id_text:
                print(f"Match Found: {installation_id_text} in Row {i}")
                # Locate the 'Accept' button in the same row
                accept_button = row.find_element(By.XPATH, ".//td[9]/a[@id='btnAccept']")
                print("Accept button found, clicking...")
                accept_button.click()
                break
        except Exception as e:
            print(f"Error processing row {i}: {e}")


def test_try_automation_publish_fall(driver, data_iteration_from, data_iteration_to, framework_variables):

    print("Inside automation_publish_fall_cmd function...")
    # VM's IP address or hostname where Mosquitto is running
    broker_address = "10.11.2.182"  # Replace with the actual IP or hostname of the VM
    port = 42244  # Custom port configured for Mosquitto on the VM

    # Credentials
    username = "appmgr"
    password = "root123"

    # Current Time Value in Epoch
    current_time_unix = int(time.time())
    fall_time_unix = current_time_unix + 120


    # Define the topic and message
    topic = "cmd/alarm/18/95"
    message = {"AS":1,"II":"2008","NI":"95","PI":18,"T":2,"TS":fall_time_unix,"V":current_time_unix}


    # Convert message to JSON
    message_json = json.dumps(message)


    # Define MQTT client and configure credentials
    client = mqtt.Client()

    # Set SSL/TLS version
    #client.tls_set(tls_version=ssl.PROTOCOL_TLSv1_2)
    #client.tls_insecure_set(True)  # Disable certificate verification

    client.username_pw_set(username, password)
    print(f"client {client}")
    print(f"client_username_pw_set {client.username_pw_set(username, password)}")

    # Connect to broker
    client.connect(broker_address, port)
    print(f"client_connect {client.connect(broker_address, port)}")

    # Publish message to the topic
    client.publish(topic, message_json, qos=1, retain=True)
    print(f"client_publish {client.publish(topic, message_json)}")

    # Disconnect from the broker
    client.disconnect()


def automation_publish_fall_command(driver, data_iteration_from, data_iteration_to, framework_variables):
    try:
        # Create an SSH client
        ssh_client = paramiko.SSHClient()

        # Automatically add the host's key if missing
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        fall_time_unix = int(time.time())
        current_time_unix = fall_time_unix - 120

        hostname = '10.11.2.212'
        port = 42115
        username = 'user'
        password ='root123'
        #command = f"""mosquitto_pub -h  broker.beanbag.online -p 42244 -t '2013/cmd/alarm/18/238' -m '{"AS":1,"II":"2008","NI":"238","PI":18,"T":2,"TS":{fall_time_unix},"V":{current_time_unix}}' -u test -P secret --capath  '/etc/ssl/certs' -I myclientid123"""
        command = """mosquitto_pub -h broker.beanbag.online -p 42244 -t '2013/cmd/alarm/18/238' -m '{{"AS":1,"II":"2008","NI":"238","PI":18,"T":2,"TS":{},"V":{}}}' -u test -P secret --capath '/etc/ssl/certs' -I myclientid123""".format(
            fall_time_unix, current_time_unix)

        # Connect to the SSH server
        ssh_client.connect(hostname, port, username, password)

        # Execute a command on the remote server
        stdin, stdout, stderr = ssh_client.exec_command(command)

        # Print the command output
        print("Command output:")
        print(stdout.read().decode())

        # Close the SSH connection
        ssh_client.close()

    except Exception as e:
        print(f"Failed to connect to hostname. Error: {e}")

def snooze_verification(driver, data_iteration_from, data_iteration_to, framework_variables):
    followup_trigger= driver.find_element(By.XPATH, "//tr[1]/td[4]")

    # Extract the text content of the WebElement
    followup_text = followup_trigger.text

    # Split the string at 'at'
    split_string = followup_text.split('at')

    # Store the first part before 'at' in a variable
    first_part = split_string[0].strip()

    # Output the result
    print("hey"+first_part+"Hey")

    if first_part == "Call center user have snoozed for 15 minutes":
        result = [framework_variables['TestScenario'],
                  framework_variables['TestID'],
                  framework_variables['row'],
                  framework_variables['datarow'],
                  "Pass",
                  first_part,
                  "Call center user have snoozed for 15 minutes",
                  "framework_variables['execution_time']"]
        custom_result_data.append(result)
    else:
        result = [framework_variables['TestScenario'],
                  framework_variables['TestID'],
                  framework_variables['row'],
                  framework_variables['datarow'],
                  "Fail",
                  first_part,
                  "Call center user have snoozed for 15 minutes",
                  "framework_variables['execution_time']"]
        custom_result_data.append(result)


def Alert_On_Support_portal_Click_Detail(driver, data_iteration_from, data_iteration_to, framework_variables):
    # Locate all rows of the table, including header
    rows = driver.find_elements(By.XPATH, "//table[@id='tblAlerts']/tbody/tr")
    print("Number of Rows: ", len(rows))

    # Iterate over each row, skipping the first row (assuming it's the header)
    for i, row in enumerate(rows, start=1):
        try:
            # Skip the header row if it contains <th> elements instead of <td>
            if row.find_elements(By.XPATH, ".//th"):
                print(f"Skipping header row {i}")
                continue

            # Locate the InstallationId element within the current row
            installation_id_element = row.find_element(By.XPATH, ".//td[2]")

            # Extract the text from the InstallationId cell
            installation_id_text = installation_id_element.text.strip()
            print(f"Row {i} Installation ID: {installation_id_text}")

            # Compare the InstallationId with the dataValue from framework_variables
            if str(framework_variables['dataValue']) == installation_id_text:
                print(f"Match Found: {installation_id_text} in Row {i}")
                # Locate the 'Accept' button in the same row
                accept_button = row.find_element(By.XPATH, ".//td[10]/a[@id='btnDetail']")
                print("Detail button found, clicking...")
                accept_button.click()
                break
        except Exception as e:
            print(f"Error processing row {i}: {e}")


def automation_publish_motion_sensor_command(driver, data_iteration_from, data_iteration_to, framework_variables):
    try:
        # Create an SSH client
        ssh_client = paramiko.SSHClient()

        # Automatically add the host's key if missing
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        current_time_unix = int(time.time())
        print(f" Current Time : {current_time_unix}")

        hostname = '10.11.2.212'
        port = 42117
        username = 'user'
        password ='root123'

        #command = """mosquitto_pub -h broker.beanbag.online -p 42244 -t '2013/evt/system/house/occupied' -m '{{"installation_id":"2008","node":"97","timestamp":{},"value":0}}' -u test -P secret --capath '/etc/ssl/certs' -I myclientid123""".format(current_time_unix)

        """command = (
                "mosquitto_pub -h broker.beanbag.online -p 42244 -t '2013/evt/system/house/occupied' "
                "-m '{\"installation_id\":\"2008\",\"node\":\"97\",\"timestamp\":%d,\"value\":0}' "
                "-u test -P secret --capath '/etc/ssl/certs' -I myclientid123" % current_time_unix
        )"""

        command = f"""mosquitto_pub -h broker.beanbag.online -p 42115 -t '2010/evt/system/house/occupied' -m '{{"installation_id":"2006","node":"43","timestamp":{current_time_unix},"value":0}}' -u test -P secret --capath '/etc/ssl/certs' -I myclientid123 -r"""

        print(f" Command : {command}")

        # Connect to the SSH server
        ssh_client.connect(hostname, port, username, password)


        # Execute a command on the remote server
        stdin, stdout, stderr = ssh_client.exec_command(command)

        # Print the command output
        print("Command output:")
        print(stdout.read().decode())

        # Close the SSH connection
        ssh_client.close()

    except Exception as e:
        print(f"Failed to connect to hostname. Error: {e}")

def automation_publish_sleep_command(driver, data_iteration_from, data_iteration_to, framework_variables):
    try:
        # Create an SSH client
        ssh_client = paramiko.SSHClient()

        # Automatically add the host's key if missing
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        current_time_unix = int(time.time())
        local_start_time_unix = current_time_unix - 7500
        local_end_time_unix =  current_time_unix - 300

        print(f" Current Time : {current_time_unix}")
        print(f" Start Time : {local_start_time_unix}")
        print(f" End Time : {local_end_time_unix}")

        hostname = '10.11.2.212'
        port = 42117
        username = 'user'
        password = 'root123'

        command = """mosquitto_pub -h broker.beanbag.online -p 42115 -t '2010/cmd/sleep/update' -m '{{"Duration":120,"InstallationId":"2006","Interrupt":2,"Predicted":0,"date":{},"local_end_time":{},"local_start_time":{}}}' -u test -P secret --capath '/etc/ssl/certs' -I myclientid123""".format(current_time_unix, local_end_time_unix, local_start_time_unix)

        print(f" Command : {command}")

        # Connect to the SSH server
        ssh_client.connect(hostname, port, username, password)


        # Execute a command on the remote server
        stdin, stdout, stderr = ssh_client.exec_command(command)

        # Print the command output
        print("Command output:")
        print(stdout.read().decode())

        # Close the SSH connection
        ssh_client.close()

    except Exception as e:
        print(f"Failed to connect to hostname. Error: {e}")

def automation_publish_away_notification_command(driver, data_iteration_from, data_iteration_to, framework_variables):
    try:
        # Create an SSH client
        ssh_client = paramiko.SSHClient()

        # Automatically add the host's key if missing
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        current_time_unix = int(time.time())

        print(f" Current Time : {current_time_unix}")

        hostname = '10.11.2.212'
        port = 42117
        username = 'user'
        password ='root123'

        command = """mosquitto_pub -h broker.beanbag.online -p 42115 -t '2010/cmd/alarm/16' -m '{{"AS":1,"II":"2006","PI":16,"T":1440,"TS":{}}}' -u test -P secret --capath '/etc/ssl/certs' -I myclientid123""".format(
            current_time_unix)

        print(f" Command : {command}")

        # Connect to the SSH server
        ssh_client.connect(hostname, port, username, password)


        # Execute a command on the remote server
        stdin, stdout, stderr = ssh_client.exec_command(command)

        # Print the command output
        print("Command output:")
        print(stdout.read().decode())

        # Close the SSH connection
        ssh_client.close()

    except Exception as e:
        print(f"Failed to connect to hostname. Error: {e}")

def automation_delete_old_notification(driver, data_iteration_from, data_iteration_to, framework_variables):
    try:
        while True:
            elements = driver.find_elements(By.XPATH,
                                            '//android.widget.ImageView[@resource-id="com.securemeters.alcarerapp:id/imgDelete"]')

            if not elements:
                print("No more items to click.")
                break

            for i in range(len(elements)):
                try:
                    # Refresh the locator for the current item
                    element = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH,
                                                    f'(//android.widget.ImageView[@resource-id="com.securemeters.alcarerapp:id/imgDelete"])[{i + 1}]'))
                    )
                    print(f"Clicking on item {i + 1}")
                    element.click()
                    break  # Exit after clicking the current item

                except Exception as e:
                    print(f"Failed to click on item {i + 1}. Error: {e}")

    except Exception as e:
        print(f"An error occurred: {e}")


def automation_rule_sync_fall_value_verification_uk(driver, data_iteration_from, data_iteration_to, framework_variables):

    from selenium import webdriver
    from selenium.webdriver.common.by import By

    text_box_id = "falldetection_fld_ptime_name"

    # Locate the text box (using ID as an example)
    text_box = driver.find_element(By.ID, text_box_id)

    # Fetch the value from the text box using get_attribute() (correct method)
    fetched_value = text_box.get_attribute("value")

    # Define the expected value
    expected_value = "2"

    if fetched_value == expected_value:
        print("Fall Duration Value matches!")
        result = [framework_variables['TestScenario'],
                  framework_variables['TestID'],
                  framework_variables['row'],
                  framework_variables['datarow'],
                  "Pass",
                  "2",
                  "2",
                  "framework_variables['execution_time']"]
        custom_result_data.append(result)
    else:
        print(f"Value does not match! Fetched: {fetched_value}, Expected: {expected_value}")
        result = [framework_variables['TestScenario'],
                  framework_variables['TestID'],
                  framework_variables['row'],
                  framework_variables['datarow'],
                  "Fail",
                  "framework_variables['actualData']",
                  "framework_variables['dataToVerify']",
                  "framework_variables['execution_time']"]
        custom_result_data.append(result)

def automation_rule_sync_fall_value_verification(driver, data_iteration_from, data_iteration_to,
                                                         framework_variables):

            from selenium import webdriver
            from selenium.webdriver.common.by import By

            text_box_id = "ptime"

            # Locate the text box (using ID as an example)
            text_box = driver.find_element(By.ID, text_box_id)

            # Fetch the value from the text box using get_attribute() (correct method)
            fetched_value = text_box.get_attribute("value")

            # Define the expected value
            expected_value = "2"

            if fetched_value == expected_value:
                print("Fall Duration Value matches!")
                result = [framework_variables['TestScenario'],
                          framework_variables['TestID'],
                          framework_variables['row'],
                          framework_variables['datarow'],
                          "Pass",
                          "2",
                          "2",
                          "framework_variables['execution_time']"]
                custom_result_data.append(result)
            else:
                print(f"Value does not match! Fetched: {fetched_value}, Expected: {expected_value}")
                result = [framework_variables['TestScenario'],
                          framework_variables['TestID'],
                          framework_variables['row'],
                          framework_variables['datarow'],
                          "Fail",
                          "framework_variables['actualData']",
                          "framework_variables['dataToVerify']",
                          "framework_variables['execution_time']"]
                custom_result_data.append(result)

def automation_rule_date_compare_ind(driver, data_iteration_from, data_iteration_to, framework_variables):
    from datetime import datetime

    current_date = datetime.now().date()

    # Locate the element using the XPath
    element = driver.find_element(By.XPATH, "//tbody/tr[2]/td[8]")

    # Get the text from the located element
    rule_time_stamp = element.text

    # Split the timestamp by space
    rule_date, rule_time = rule_time_stamp.split(" ")

    print("Date:", rule_date)
    print("Time:", rule_time)

    # Convert strings to datetime objects
    date2 = datetime.strptime(rule_date, "%Y-%m-%d").date()

    if current_date < date2:
        print(f"{current_date} is earlier than {date2}")
        result = [framework_variables['TestScenario'],
                  framework_variables['TestID'],
                  framework_variables['row'],
                  framework_variables['datarow'],
                  "Fail",
                  current_date,
                  date2,
                  "framework_variables['execution_time']"]
        custom_result_data.append(result)
    elif current_date > date2:
        print(f"{current_date} is later than {date2}")
        result = [framework_variables['TestScenario'],
                  framework_variables['TestID'],
                  framework_variables['row'],
                  framework_variables['datarow'],
                  "Fail",
                  current_date,
                  date2,
                  "framework_variables['execution_time']"]
        custom_result_data.append(result)
    else:
        print(f"{current_date} is the same as {date2}")
        result = [framework_variables['TestScenario'],
                  framework_variables['TestID'],
                  framework_variables['row'],
                  framework_variables['datarow'],
                  "Pass",
                  current_date,
                  date2,
                  "framework_variables['execution_time']"]
        custom_result_data.append(result)

def automation_rule_date_compare_uk(driver, data_iteration_from, data_iteration_to, framework_variables):
    from datetime import datetime

    # Get the current date
    current_date = datetime.now().date()

    # Locate the element using the XPath
    element = driver.find_element(By.XPATH, "//tbody/tr[1]/td[8]")

    # Get the text from the located element (assumed to be a timestamp in 'dd-mm-yyyy hh:mm:ss' format)
    rule_time_stamp = element.text  # e.g., "09-10-2024 11:51:19"

    # Split the timestamp by space to separate date and time (if needed)
    rule_date_time = datetime.strptime(rule_time_stamp, "%d-%m-%Y %H:%M:%S")

    # Extract the date part
    rule_date = rule_date_time.date()

    print("Rule Date:", rule_date)
    print("Current Date:", current_date)

    # Compare the current date (current_date) with the rule date (rule_date)
    if current_date < rule_date:
        print(f"{current_date} is earlier than {rule_date}")
        result = [framework_variables['TestScenario'],
                  framework_variables['TestID'],
                  framework_variables['row'],
                  framework_variables['datarow'],
                  "Fail",
                  current_date,
                  rule_date,
                  "framework_variables['execution_time']"]
        custom_result_data.append(result)
    elif current_date > rule_date:
        print(f"{current_date} is later than {rule_date}")
        result = [framework_variables['TestScenario'],
                  framework_variables['TestID'],
                  framework_variables['row'],
                  framework_variables['datarow'],
                  "Fail",
                  current_date,
                  rule_date,
                  "framework_variables['execution_time']"]
        custom_result_data.append(result)
    else:
        print(f"{current_date} is the same as {rule_date}")
        result = [framework_variables['TestScenario'],
                  framework_variables['TestID'],
                  framework_variables['row'],
                  framework_variables['datarow'],
                  "Pass",
                  current_date,
                  rule_date,
                  "framework_variables['execution_time']"]
        custom_result_data.append(result)

def automation_publish_unreachable_alert_generation_command(driver, data_iteration_from, data_iteration_to, framework_variables):
    try:
        # Create an SSH client
        ssh_client = paramiko.SSHClient()

        # Automatically add the host's key if missing
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        current_time_unix = int(time.time())

        print(f" Current Time : {current_time_unix}")

        hostname = '10.11.2.182'
        port = 22
        username = 'appmgr'
        password ='root123'

        command = """mosquitto_pub -h broker.beanbag.online -p 42244 -t '2013/cmd/alarm/24/-1' -m '{{"AS":1,"II":"2008","NI":-1,"PI":24,"T":2880,"TS":{}}}' -u test -P secret --capath '/etc/ssl/certs' -I myclientid123""".format(
            current_time_unix)

        print(f" Command : {command}")

        # Connect to the SSH server
        ssh_client.connect(hostname, port, username, password)


        # Execute a command on the remote server
        stdin, stdout, stderr = ssh_client.exec_command(command)

        # Print the command output
        print("Command output:")
        print(stdout.read().decode())

        # Close the SSH connection
        ssh_client.close()

    except Exception as e:
        print(f"Failed to connect to hostname. Error: {e}")

def automation_publish_unreachable_alert_restoration_command(driver, data_iteration_from, data_iteration_to, framework_variables):
    try:
        # Create an SSH client
        ssh_client = paramiko.SSHClient()

        # Automatically add the host's key if missing
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        current_time_unix = int(time.time())

        print(f" Current Time : {current_time_unix}")

        hostname = '10.11.2.182'
        port = 22
        username = 'appmgr'
        password ='root123'

        command = """mosquitto_pub -h broker.beanbag.online -p 42244 -t '2013/cmd/alarm/24/-1' -m '{{"AS":3,"II":"2008","NI":-1,"PI":24,"T":2880,"TS":{}}}' -u test -P secret --capath '/etc/ssl/certs' -I myclientid123""".format(
            current_time_unix)

        print(f" Command : {command}")

        # Connect to the SSH server
        ssh_client.connect(hostname, port, username, password)


        # Execute a command on the remote server
        stdin, stdout, stderr = ssh_client.exec_command(command)

        # Print the command output
        print("Command output:")
        print(stdout.read().decode())

        # Close the SSH connection
        ssh_client.close()

    except Exception as e:
        print(f"Failed to connect to hostname. Error: {e}")


def test_sql_execution(driver, data_iteration_from, data_iteration_to, framework_variables):

    # PV Database connection setting
    conn = psycopg2.connect(

        host="10.11.2.211",
        user="smartadmin",
        password="smart123",
        database="assistedliving",
        port="5432"
    )

    cursor = conn.cursor()

    try:
        # Execute the SELECT query
        select_query = "select * from accesstype"  # Modify as per your table and columns
        cursor.execute(select_query)

        # Fetch all rows from the result of the query
        rows = cursor.fetchall()

        # Display the results
        for row in rows:
            print(row)

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Close the cursor and connection
        cursor.close()
        conn.close()

def sos_i_am_okay(driver, data_iteration_from, data_iteration_to, framework_variables):
    # This Function will accept ID locator and only ID value not package Name.
    print("Inside Wait Until Element Visible Function")
    locator_value = "com.securemeters.alapp:id/tv_Positive"

    print(locator_value)
    try:
        # Initialize WebDriverWait with a 300-second timeout
        wait = WebDriverWait(driver, 300)

        # Wait until the element is visible, using the XPATH from framework_variables
        element = wait.until(EC.visibility_of_element_located((By.ID, locator_value)))

        print("Element is visible!")

        text_value = element.text
        print("Locator Text:", text_value)

        # Focus and tap using TouchAction
        actions = TouchAction(driver)
        actions.tap(element).perform()

        # Perform actions on the element if needed, e.g., click or read text
        #element.click()  # Example action

        print("Click on Element")

    except Exception as e:
        print(f"An error occurred: {e}")


def restart_appium(driver, data_iteration_from, data_iteration_to, framework_variables):
    print("Inside restart_appium function")
    batch_file_path = r'D:\Automation Framework\Regression Cycle 1 Remaining\Assisted_Living\ProjectScripts\restart_appium.bat'
    subprocess.run([batch_file_path], shell=True)

def OpenRegression(driver, data_iteration_from, data_iteration_to, framework_variables):
    print("Inside restart_appium function")
    batch_file_path = r'D:\Automation Framework\Regression Cycle 1 Remaining\Assisted_Living\ProjectScripts\OpenRegression.bat'
    subprocess.run([batch_file_path], shell=True)


def find_unclose_alert(driver, data_iteration_from, data_iteration_to, framework_variables):
    from selenium.webdriver.common.by import By
    import time

    # Locate all rows of the table, including the header
    rows = driver.find_elements(By.XPATH, "//table[@id='tblAlerts']/tbody/tr")
    print("Number of Rows: ", len(rows))

    # Iterate over each row, skipping the first row (assuming it's the header)
    for i, row in enumerate(rows, start=1):
        try:
            # Skip the header row if it contains <th> elements instead of <td>
            if row.find_elements(By.XPATH, ".//th"):
                print(f"Skipping header row {i}")
                continue

            # Locate the InstallationId element within the current row
            installation_id_element = row.find_element(By.XPATH, ".//td[2]")

            # Extract the text from the InstallationId cell
            installation_id_text = installation_id_element.text.strip()
            print(f"Row {i} Installation ID: {installation_id_text}")

            # Compare the InstallationId with the dataValue from framework_variables
            if str(framework_variables['dataValue']) == installation_id_text:
                print(f"Match Found: {installation_id_text} in Row {i}")
                try:
                    # Attempt to locate and click the 'Accept' button
                    accept_button = row.find_element(By.XPATH, ".//td[9]/a[@id='btnAccept']")
                    print("Accept button found, clicking...")
                    accept_button.click()
                    time.sleep(3)
                except Exception as e:
                    print(f"Accept button not accessible for Row {i}: {e}")
                    try:
                        # Locate and click the 'Detail' button if 'Accept' is not clickable
                        detail_button = row.find_element(By.XPATH, ".//td[10]/a[@id='btnDetail']")
                        print("Clicking Detail button instead...")
                        detail_button.click()
                    except Exception as e:
                        print(f"Detail button not accessible for Row {i}: {e}")
                break
        except Exception as e:
            print(f"Error processing row {i}: {e}")

def close_alert_mqtt(driver, data_iteration_from, data_iteration_to, framework_variables):
    # Split the string into a list
    values = framework_variables['dataValue'].split(',')

    # Assign the values to separate variables
    alert_id = values[0]
    alert_status = values[1]

    # Print the results
    print("alert_id:", alert_id)
    print("alert_status:", alert_status)

    # Check the condition
    if alert_status == 'InProgress' or alert_status == 'Pending' or alert_status == 'NoResponse':
        try:
            # Create an SSH client
            ssh_client = paramiko.SSHClient()
            # Automatically add the host's key if missing
            ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            current_time_unix = int(time.time())
            print(f" Current Time : {current_time_unix}")
            hostname = '10.11.2.212'
            port = 42117
            username = 'user'
            password ='root123'
            command = f"""mosquitto_pub -h broker.beanbag.online -p 42115 -t '2010/cmd/alarm/11' -m '{{"I":{alert_id},"II":2006,"PI":11,"AS":3,"NI":null,"TS":{current_time_unix},"D":{{"AT":2,"C":0,"R":"I am okay","DUI":6092}}}}' -u test -P secret --capath '/etc/ssl/certs' -I myclientid123"""

            print(f" Command : {command}")
            # Connect to the SSH server
            ssh_client.connect(hostname, port, username, password)
            # Execute a command on the remote server
            stdin, stdout, stderr = ssh_client.exec_command(command)
        # Print the command output
            print("Command output:")
            print(stdout.read().decode())

            # Close the SSH connection
            ssh_client.close()

        except Exception as e:
            print(f"Failed to connect to hostname. Error: {e}")

    else:
        print("Breaking out of the process")
        # If using a loop, add a `break` statement here
